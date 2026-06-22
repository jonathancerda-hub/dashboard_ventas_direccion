# app.py - Dashboard de Ventas Farmacéuticas

# Usar el almacén de certificados del SO (Windows) para TLS. Necesario en local tras
# proxies corporativos con inspección SSL (Google OAuth/Sheets fallan con certifi).
# Inofensivo en Render (si truststore no está instalado, simplemente se omite).
try:
    import truststore
    truststore.inject_into_ssl()
except Exception:
    pass

from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from dotenv import load_dotenv
from odoo_manager import OdooManager
from google_sheets_manager import GoogleSheetsManager
import os
import pandas as pd
import json
import io
import calendar
import re
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import hashlib
import pickle

load_dotenv(override=True)
app = Flask(__name__)

# --- Configuración de Supabase (Datos Históricos) ---
try:
    from supabase_manager import SupabaseManager
    supabase_manager = SupabaseManager()
    SUPABASE_ENABLED = True
    print("✅ Supabase habilitado para datos históricos")
except Exception as e:
    SUPABASE_ENABLED = False
    print(f"⚠️ Supabase no disponible: {e}")
    print("   Continuando con Odoo únicamente...")

# --- Configuración de la Clave Secreta ---
app.secret_key = os.getenv("SECRET_KEY")
if not app.secret_key:
    print("⚠️  Advertencia: La variable de entorno SECRET_KEY no está configurada.")
    print("La sesión de Flask no funcionará (ej. login, flash messages).")

# --- Lista de usuarios administradores ---
ADMIN_USERS = os.getenv("ADMIN_USERS", "").split(",")
ADMIN_USERS = [email.strip() for email in ADMIN_USERS if email.strip()]

# --- Configuración Google OAuth (login con Google) ---
GOOGLE_CLIENT_ID = os.getenv('GOOGLE_CLIENT_ID')
GOOGLE_CLIENT_SECRET = os.getenv('GOOGLE_CLIENT_SECRET')
GOOGLE_OAUTH_ENABLED = bool(GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET)
# Google añade scopes ('email','profile') además de los pedidos -> evitar que oauthlib
# lance "Scope has changed". Inofensivo en todos los entornos.
os.environ.setdefault('OAUTHLIB_RELAX_TOKEN_SCOPE', '1')
# En local el redirect es http://localhost (no https): oauthlib lo rechaza por defecto.
# Solo permitir transporte inseguro fuera de Render (en producción es https).
if os.getenv('RENDER', 'false').lower() != 'true':
    os.environ.setdefault('OAUTHLIB_INSECURE_TRANSPORT', '1')
OAUTH_SCOPES = [
    'openid',
    'https://www.googleapis.com/auth/userinfo.email',
    'https://www.googleapis.com/auth/userinfo.profile',
]
if not GOOGLE_OAUTH_ENABLED:
    print("⚠️  Google OAuth no configurado (faltan GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET en .env)")


def get_allowed_emails():
    """Lista blanca de correos autorizados (env ALLOWED_USERS o allowed_users.json)."""
    allowed_emails_env = os.getenv('ALLOWED_USERS')
    if allowed_emails_env:
        return [e.strip() for e in allowed_emails_env.split(',') if e.strip()]
    allowed_users_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'allowed_users.json')
    with open(allowed_users_path, 'r') as f:
        return json.load(f).get('allowed_emails', [])


def _build_oauth_flow(state=None):
    """Crea el Flow de OAuth de Google usando las credenciales del .env."""
    from google_auth_oauthlib.flow import Flow
    client_config = {
        "web": {
            "client_id": GOOGLE_CLIENT_ID,
            "client_secret": GOOGLE_CLIENT_SECRET,
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
        }
    }
    redirect_uri = url_for('oauth2callback', _external=True,
                           _scheme='https' if IS_RENDER else 'http')
    return Flow.from_client_config(client_config, scopes=OAUTH_SCOPES,
                                   state=state, redirect_uri=redirect_uri)

# --- Configuración de Templates ---
# En producción (Render): habilitar caché de templates para evitar timeouts
# En desarrollo: deshabilitar caché para ver cambios inmediatos
IS_RENDER = os.getenv('RENDER', 'false').lower() == 'true'
if IS_RENDER:
    app.config['TEMPLATES_AUTO_RELOAD'] = False  # Habilitar caché en producción
    # Habilitar bytecode cache de Jinja2 para compilación rápida
    from jinja2 import FileSystemBytecodeCache
    bytecode_cache_dir = os.path.join(os.path.dirname(__file__), '__pycache__', 'jinja2_cache')
    os.makedirs(bytecode_cache_dir, exist_ok=True)
    app.jinja_env.bytecode_cache = FileSystemBytecodeCache(bytecode_cache_dir)
    print("✅ Caché de templates Jinja2 habilitado (producción)")
else:
    app.config['TEMPLATES_AUTO_RELOAD'] = True  # Recargar templates en desarrollo
    print("💻 Auto-reload de templates habilitado (desarrollo)")

app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# --- Sistema de Caché para Datos de Meses ---
CACHE_DIR = os.path.join(os.path.dirname(__file__), '__pycache__', 'dashboard_cache')
os.makedirs(CACHE_DIR, exist_ok=True)
CACHE_TTL_CURRENT_MONTH = 1800 # 30 minutos

# --- Límite de líneas de ventas (para evitar sobrecarga en Render) ---
try:
    SALES_LIMIT = int(os.getenv('SALES_LIMIT', '5000'))
except Exception:
    SALES_LIMIT = 5000

def get_cache_key(año, mes):
    """Genera una clave única para el caché basada en año y mes."""
    return f"dashboard_data_{año}_{mes:02d}"

def is_current_month(año, mes):
    """Verifica si el mes solicitado es el mes actual."""
    hoy = datetime.now()
    return año == hoy.year and mes == hoy.month

def get_cached_data(año, mes):
    """Obtiene datos del caché si existen y son válidos."""
    cache_key = get_cache_key(año, mes)
    cache_file = os.path.join(CACHE_DIR, f"{cache_key}.pkl")
    
    if not os.path.exists(cache_file):
        return None
    
    try:
        with open(cache_file, 'rb') as f:
            timestamp, cached_data = pickle.load(f)
        
        # Para el mes actual, verificar si el caché ha expirado
        if is_current_month(año, mes):
            age = (datetime.now() - timestamp).total_seconds()
            if age > CACHE_TTL_CURRENT_MONTH:
                print(f"🕒 Caché para mes actual expirado (antigüedad: {age:.0f}s). Se necesita refrescar.")
                return None
            print(f"⚡️ Datos cargados desde caché para MES ACTUAL (antigüedad: {age:.0f}s)")
            return cached_data

        # Para meses pasados, el caché es válido indefinidamente
        print(f"✅ Datos cargados desde caché para mes pasado ({año}-{mes:02d})")
        return cached_data
    except Exception as e:
        print(f"⚠️ Error al leer caché ({e.__class__.__name__}): {e}. Se tratará como sin caché.")
        return None

def save_to_cache(año, mes, data):
    """Guarda datos en el caché con un timestamp."""
    cache_key = get_cache_key(año, mes)
    cache_file = os.path.join(CACHE_DIR, f"{cache_key}.pkl")
    
    try:
        # Guardar siempre con el timestamp actual
        data_to_cache = (datetime.now(), data)
        with open(cache_file, 'wb') as f:
            pickle.dump(data_to_cache, f)
        print(f"💾 Datos guardados en caché para {año}-{mes:02d}")
    except Exception as e:
        print(f"⚠️ Error al guardar en caché: {e}")

# --- Inicialización de Managers ---
try:
    data_manager = OdooManager()
except Exception as e:
    print(f"⚠️ No se pudo inicializar OdooManager: {e}. Continuando en modo offline.")
    # Crear un stub mínimo con las funciones usadas en la app para evitar fallos
    class _StubManager:
        def get_filter_options(self):
            return {'lineas': [], 'clients': []}
        def get_sales_lines(self, *args, **kwargs):
            return []
        def get_all_sellers(self):
            return []
        def get_commercial_lines_stacked_data(self, *args, **kwargs):
            return {'yAxis': [], 'series': [], 'legend': []}
    data_manager = _StubManager()
gs_manager = GoogleSheetsManager(
    credentials_file='credentials.json',
    sheet_name=os.getenv('GOOGLE_SHEET_NAME')
)

# --- Funciones Auxiliares ---

def get_meses_del_año(año):
    """Genera una lista de meses para un año específico."""
    meses_nombres = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    meses_disponibles = []
    for i in range(1, 13):
        mes_key = f"{año}-{i:02d}"
        mes_nombre = f"{meses_nombres[i-1]} {año}"
        meses_disponibles.append({'key': mes_key, 'nombre': mes_nombre})
    return meses_disponibles

def get_data_source(año: int, mes: int = None):
    """
    Determina de dónde obtener los datos según el año (y opcionalmente el mes).

    Reglas:
    - Años históricos (≤2025): Supabase si hay datos.
    - Año actual / futuro (2026+): el MES EN CURSO y meses futuros van a Odoo en vivo;
      los meses CERRADOS ya migrados a Supabase (vía migrar_mes.py) se leen de Supabase.
    - Si no se pasa `mes`, se mantiene el comportamiento por año (compatibilidad):
      2026+ -> Odoo.

    Args:
        año: Año a consultar
        mes: Mes (1-12) opcional. Necesario para enrutar meses cerrados de 2026+.

    Returns:
        'supabase' o 'odoo'
    """
    # Años históricos (2025 y anteriores) en Supabase
    if SUPABASE_ENABLED and año <= 2025:
        has_data = supabase_manager.is_year_in_supabase(año)
        print(f"🔍 Verificando año {año} en Supabase: {'✅ Encontrado' if has_data else '❌ No encontrado'}")
        if has_data:
            return 'supabase'

    # Año actual/futuro con mes concreto: meses cerrados migrados -> Supabase
    if SUPABASE_ENABLED and mes is not None and año >= 2026:
        hoy = datetime.now()
        es_mes_en_curso = (año > hoy.year) or (año == hoy.year and mes >= hoy.month)
        if not es_mes_en_curso and supabase_manager.is_month_in_supabase(año, mes):
            print(f"🔍 Mes cerrado {año}-{mes:02d} encontrado en Supabase ✅")
            return 'supabase'

    print(f"🔄 Usando Odoo para año {año}" + (f" mes {mes:02d}" if mes else ""))
    return 'odoo'

def load_metas(año: int):
    """
    Carga las metas desde la fuente correcta según el año
    
    Args:
        año: Año para cargar las metas
    
    Returns:
        Dict con formato: {
            'mes_key': {
                'metas': {'linea': valor, ...},
                'metas_ipn': {'linea': valor, ...},
                'total': float,
                'total_ipn': float
            }
        }
    """
    # Para 2025 y posteriores, usar Supabase (tabla metas_ventas_2026 contiene 2025 y 2026)
    if año >= 2025 and SUPABASE_ENABLED:
        print(f"📊 Cargando metas del {año} desde Supabase...")
        return supabase_manager.read_metas_from_supabase(año=año)
    
    # Para años anteriores a 2025, usar Google Sheets
    print(f"📊 Cargando metas desde Google Sheets...")
    return gs_manager.read_metas_por_linea()

def normalizar_linea_comercial(nombre_linea):
    """
    Normaliza nombres de líneas comerciales agrupando GENVET y MARCA BLANCA como TERCEROS.
    
    Ejemplos:
    - GENVET → TERCEROS
    - MARCA BLANCA → TERCEROS
    - GENVET PERÚ → TERCEROS
    - PETMEDICA → PETMEDICA (sin cambios)
    """
    if not nombre_linea:
        return nombre_linea
    
    nombre_upper = nombre_linea.upper().strip()
    
    # Agrupar GENVET y MARCA BLANCA como TERCEROS
    if 'GENVET' in nombre_upper or 'MARCA BLANCA' in nombre_upper:
        return 'TERCEROS'
    
    return nombre_linea.upper().strip()

ATREVIA_TAMANOS_PRESENTACIONES_ORDENADOS = sorted([
    'MEDIUM', 'LARGE', 'SMALL', 'MINI', 'EXTRA LARGE', 'XL', 'L', 'M', 'S',
    'SPOT ON MEDIUM', 'SPOT ON LARGE', 'SPOT ON SMALL', 'SPOT ON MINI',
    'CATS SPOT ON MEDIUM', 'CATS SPOT ON LARGE', 'CATS SPOT ON SMALL', 'CATS SPOT ON MINI',
    'SPOT ON'
], key=len, reverse=True)

def limpiar_nombre_atrevia(nombre_producto):
    """
    Limpia los nombres de productos ATREVIA eliminando indicadores de tamaño/presentación.
    
    Ejemplos:
    - ATREVIA ONE MEDIUM → ATREVIA ONE
    - ATREVIA XR LARGE → ATREVIA XR  
    - ATREVIA 360° MEDIUM → ATREVIA 360°
    - ATREVIA TRIO CATS SPOT ON MEDIUM → ATREVIA TRIO CATS
    """
    if not nombre_producto or 'ATREVIA' not in nombre_producto.upper():
        return nombre_producto
    
    nombre_limpio = nombre_producto.strip()
    
    # Procesar solo si contiene ATREVIA
    if 'ATREVIA' in nombre_limpio.upper():
        for tamano in ATREVIA_TAMANOS_PRESENTACIONES_ORDENADOS:
            # Buscar y eliminar el tamaño/presentación al final del nombre
            if nombre_limpio.upper().endswith(' ' + tamano):
                nombre_limpio = nombre_limpio[:-(len(tamano) + 1)].strip()
                break
    
    return nombre_limpio

def extraer_nombre_relacional(valor_relacional, default=''):
    """Extrae el nombre de un campo relacional de Odoo con formato [id, nombre]."""
    if isinstance(valor_relacional, list) and len(valor_relacional) > 1:
        return valor_relacional[1]
    return default

@app.route('/login')
def login():
    # Si ya hay sesión, ir directo al dashboard
    if 'username' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html', google_oauth_enabled=GOOGLE_OAUTH_ENABLED)


@app.route('/login/google')
def login_google():
    """Inicia el flujo OAuth: redirige al consentimiento de Google."""
    if not GOOGLE_OAUTH_ENABLED:
        flash('El inicio de sesión con Google no está configurado.', 'danger')
        return redirect(url_for('login'))
    flow = _build_oauth_flow()
    auth_url, state = flow.authorization_url(
        access_type='offline',
        include_granted_scopes='true',
        prompt='select_account',
    )
    session['oauth_state'] = state
    return redirect(auth_url)


@app.route('/oauth2callback')
def oauth2callback():
    """Callback de Google: valida el token, comprueba lista blanca y crea sesión."""
    if not GOOGLE_OAUTH_ENABLED:
        flash('El inicio de sesión con Google no está configurado.', 'danger')
        return redirect(url_for('login'))

    # Verificación de state (protección CSRF)
    state = session.get('oauth_state')
    if not state or request.args.get('state') != state:
        flash('Sesión de login inválida o expirada. Intenta de nuevo.', 'danger')
        return redirect(url_for('login'))

    # Si el usuario canceló el consentimiento
    if request.args.get('error'):
        flash('Inicio de sesión con Google cancelado.', 'warning')
        return redirect(url_for('login'))

    try:
        flow = _build_oauth_flow(state=state)
        flow.fetch_token(authorization_response=request.url)
        credentials = flow.credentials

        # Validar el id_token y extraer el email verificado
        from google.oauth2 import id_token as google_id_token
        from google.auth.transport import requests as google_requests
        id_info = google_id_token.verify_oauth2_token(
            credentials.id_token, google_requests.Request(), GOOGLE_CLIENT_ID,
            clock_skew_in_seconds=10,
        )

        email = (id_info.get('email') or '').strip()
        email_verified = id_info.get('email_verified', False)
        nombre = id_info.get('name', email)

        if not email or not email_verified:
            flash('No se pudo verificar tu correo de Google.', 'danger')
            return redirect(url_for('login'))

        # --- Verificación de Lista Blanca ---
        try:
            allowed_emails = get_allowed_emails()
        except FileNotFoundError:
            flash('Error de configuración: archivo de usuarios permitidos no encontrado.', 'danger')
            return redirect(url_for('login'))

        # Comparación sin distinción de mayúsculas/minúsculas
        allowed_lower = [e.lower() for e in allowed_emails]
        if email.lower() in allowed_lower:
            session['username'] = email
            session['user_name'] = nombre
            session.pop('oauth_state', None)
            flash('¡Inicio de sesión exitoso!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash(f'La cuenta {email} no tiene permiso para acceder a esta aplicación.', 'warning')
            return redirect(url_for('login'))

    except Exception as e:
        print(f"❌ Error en OAuth callback: {e}")
        flash('Error al iniciar sesión con Google. Intenta de nuevo.', 'danger')
        return redirect(url_for('login'))


@app.route('/logout')
def logout():
    session.clear()
    flash('Has cerrado sesión correctamente.', 'info')
    return redirect(url_for('login'))

@app.route('/')
def index():
    # Redirigir la ruta raíz al dashboard
    return redirect(url_for('dashboard'))


@app.route('/api/tendencia', methods=['GET'])
def api_tendencia():
    """API para obtener datos de tendencia histórica de un año específico"""
    if 'username' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        año = int(request.args.get('año', datetime.now().year))
        
        print(f"📊 API Tendencia: Solicitando datos para año {año}")
        
        # Obtener datos de tendencia (optimizado para Render Free)
        fecha_inicio = f"{año}-01-01"
        
        # Determinar fuente de datos (probar enero: usa Supabase si el año está migrado)
        data_source = get_data_source(año, 1)
        
        # Optimización: Si es año actual con Odoo, solo hasta hoy
        if data_source == 'odoo' and año == datetime.now().year:
            fecha_fin = datetime.now().strftime('%Y-%m-%d')
            print(f"📅 API Tendencia: Año actual, consultando solo hasta {fecha_fin}")
        else:
            fecha_fin = f"{año}-12-31"
        print(f"📊 Fuente de datos para {año}: {data_source}")
        
        # Obtener resumen mensual
        if data_source == 'supabase':
            # Para Supabase, usar get_sales_by_month que ya suma correctamente sin filtros restrictivos
            print(f"🔥 API TENDENCIA: Cargando datos de Supabase para {año}")
            resumen_mensual = supabase_manager.get_sales_by_month(fecha_inicio, fecha_fin)
            print(f"✅ API TENDENCIA: Resumen mensual obtenido - {len(resumen_mensual)} meses con datos")
            if 'enero 2025' in resumen_mensual:
                print(f"   📊 Enero 2025: S/ {resumen_mensual['enero 2025']:,.2f}")
        else:
            resumen_mensual = data_manager.get_sales_summary_by_month(fecha_inicio, fecha_fin)
        
        # Obtener metas del año desde la fuente correcta (Google Sheets o Supabase)
        metas_historicas = load_metas(año)
        
        # Construir array de 12 meses
        tendencia = []
        meses_es = {
            1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio',
            7: 'julio', 8: 'agosto', 9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
        }
        
        for mes_num in range(1, 13):
            fecha_mes = datetime(año, mes_num, 1)
            mes_key = f"{año}-{mes_num:02d}"
            
            # Buscar venta del mes (formato esperado: "enero 2025", "febrero 2025", etc.)
            label_busqueda = f"{meses_es[mes_num]} {año}"
            venta_mes = resumen_mensual.get(label_busqueda, 0)
            
            # DEBUG temporal
            if año == 2025 and mes_num == 1:
                print(f"🔍 API /tendencia - Buscando: '{label_busqueda}'")
                print(f"🔍 API /tendencia - Encontrado: {venta_mes}")
                print(f"🔍 API /tendencia - Keys en resumen_mensual: {list(resumen_mensual.keys())[:3]}")
            
            # Buscar meta del mes
            try:
                meta_key = f"{año}-{mes_num:02d}"
                metas_mes_data = metas_historicas.get(meta_key, {}).get('metas', {})
                meta_mes = sum(metas_mes_data.values())
            except:
                meta_mes = 0
            
            tendencia.append({
                'mes': mes_key,
                'mes_nombre': fecha_mes.strftime('%b %Y'),
                'venta': venta_mes,
                'meta': meta_mes,
                'cumplimiento': (venta_mes / meta_mes * 100) if meta_mes > 0 else 0
            })
        
        print(f"✅ API Tendencia: {len(tendencia)} meses procesados para {año}")
        
        response = jsonify({
            'success': True,
            'año': año,
            'tendencia': tendencia,
            'fuente': data_source
        })
        # Agregar headers de no-caché
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
        
    except Exception as e:
        print(f"❌ Error en API tendencia: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/rfm-canal', methods=['GET'])
def api_rfm_canal():
    """API para obtener análisis RFM filtrado por canal"""
    if 'username' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        canal_filtro = request.args.get('canal', 'TODOS').upper()
        año = int(request.args.get('año', datetime.now().year))
        mes = int(request.args.get('mes', datetime.now().month))
        
        print(f"📊 API RFM Canal: Solicitando datos para canal={canal_filtro}, año={año}, mes={mes}")
        
        # Obtener datos en caché o calcular
        cached_data = get_cached_data(año, mes)
        
        if cached_data and 'clientes_rfm' in cached_data and 'segmentos_por_canal' in cached_data:
            clientes_rfm = cached_data['clientes_rfm']
            segmentos_por_canal = cached_data['segmentos_por_canal']
            
            # Filtrar según canal seleccionado
            if canal_filtro == 'TODOS':
                clientes_filtrados = clientes_rfm
                segmentos_filtrados = segmentos_por_canal.get('TODOS', {})
            elif canal_filtro in ['DIGITAL', 'NACIONAL', 'OTROS']:
                clientes_filtrados = [c for c in clientes_rfm if c.get('canal', 'SIN CANAL') == canal_filtro or (canal_filtro == 'OTROS' and c.get('canal', 'SIN CANAL') not in ['DIGITAL', 'NACIONAL'])]
                segmentos_filtrados = segmentos_por_canal.get(canal_filtro, {})
            else:
                clientes_filtrados = clientes_rfm
                segmentos_filtrados = segmentos_por_canal.get('TODOS', {})
            
            # Ordenar y limitar a top 100
            clientes_filtrados = sorted(clientes_filtrados, key=lambda x: x['monetary'], reverse=True)[:100]
            
            print(f"✅ API RFM Canal: {len(clientes_filtrados)} clientes filtrados para {canal_filtro}")
            
            return jsonify({
                'success': True,
                'canal': canal_filtro,
                'clientes': clientes_filtrados,
                'segmentos': segmentos_filtrados,
                'total_clientes': len(clientes_filtrados)
            })
        else:
            return jsonify({'error': 'Datos no disponibles en caché. Recargue el dashboard.'}), 404
        
    except Exception as e:
        print(f"❌ Error en API RFM Canal: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def generar_datos_ventas_mes(año_para_grafico, data_source, fecha_actual, sales_data_anual_override=None):
    """
    Genera los datos del gráfico de ventas por mes con filtros farmacéuticos.
    Esta función se ejecuta siempre, incluso cuando hay caché.
    Lee directamente de Supabase o Odoo según el año.
    """
    meses_español = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    
    ventas_detalladas_por_mes = []
    filtros_disponibles = {
        'lineas_comerciales': set(),
        'categorias': set(),
        'ciclos_vida': set(),
        'vias_administracion': set(),
        'clasificaciones': set(),
        'formas_farmaceuticas': set(),
        'lineas_produccion': set()
    }
    
    try:
        normalizar_linea = normalizar_linea_comercial
        extraer_rel = extraer_nombre_relacional
        meses_get = meses_español.get
        append_registro = ventas_detalladas_por_mes.append
        filtros_lineas = filtros_disponibles['lineas_comerciales']
        filtros_categorias = filtros_disponibles['categorias']
        filtros_ciclos = filtros_disponibles['ciclos_vida']
        filtros_vias = filtros_disponibles['vias_administracion']
        filtros_clasif = filtros_disponibles['clasificaciones']
        filtros_formas = filtros_disponibles['formas_farmaceuticas']
        filtros_prod = filtros_disponibles['lineas_produccion']
        lineas_invalidas = {'Sin Línea', 'NINGUNO', ''}

        fecha_inicio_anual = f"{año_para_grafico}-01-01"
        
        # Optimización Render Free: Solo consultar hasta el mes actual, no todo el año
        # Esto reduce drasticamente el tiempo de consulta a Odoo (de 300s a 30s)
        if data_source == 'odoo' and año_para_grafico == fecha_actual.year:
            # Año actual: solo consultar hasta fin del mes actual
            fecha_fin_anual = f"{año_para_grafico}-{fecha_actual.month:02d}-{fecha_actual.day:02d}"
            print(f"📅 Año actual: Cargando solo hasta hoy {fecha_fin_anual} (optimización Render)")
        else:
            # Año histórico en Supabase: consultar todo el año
            fecha_fin_anual = f"{año_para_grafico}-12-31"
            print(f"📅 Cargando ventas anuales: {fecha_inicio_anual} a {fecha_fin_anual}")
        
        if sales_data_anual_override is not None:
            sales_data_anual = sales_data_anual_override
            print(f"⚡ Ventas anuales reutilizadas desde contexto: {len(sales_data_anual)} registros")
        elif data_source == 'supabase':
            sales_data_anual = supabase_manager.get_dashboard_data(fecha_inicio_anual, fecha_fin_anual)
        else:
            sales_data_anual = data_manager.get_sales_lines(
                date_from=fecha_inicio_anual,
                date_to=fecha_fin_anual,
                limit=50000
            )
        
        print(f"📊 Ventas anuales obtenidas: {len(sales_data_anual)} registros")
        
        for sale in sales_data_anual:
            # Manejar diferentes formatos de fecha según fuente
            fecha_str = sale.get('invoice_date') or sale.get('date_order', '')
            if not fecha_str:
                continue
                
            try:
                mes_num = int(fecha_str[5:7])
                mes_nombre = meses_get(mes_num, '')
                
                if not mes_nombre:
                    continue
                
                # Extraer campos (tanto Supabase como Odoo tienen la misma estructura ahora)
                if data_source == 'supabase':
                    # Supabase: campos ya vienen como strings simples
                    linea_comercial = sale.get('commercial_line_name', 'Sin Línea')
                    
                    # Filtrar VENTA INTERNACIONAL (igual que KPI Venta)
                    if linea_comercial and 'VENTA INTERNACIONAL' in str(linea_comercial).upper():
                        continue
                    
                    # Filtrar ventas sin línea comercial válida (igual que KPI Venta)
                    if not linea_comercial or linea_comercial in lineas_invalidas:
                        continue
                    
                    if linea_comercial and linea_comercial != 'Sin Línea':
                        linea_comercial = normalizar_linea(linea_comercial)
                    
                    categoria = sale.get('category_name', 'Sin Categoría')
                    via_administracion = sale.get('administration_way_name', 'No Definido')
                    clasificacion = sale.get('pharmacological_classification_name', 'No Definido')
                    forma_farmaceutica = sale.get('pharmaceutical_forms_name', 'No Definido')
                    linea_produccion = sale.get('production_line_name', 'No Definido')
                    ciclo_vida = sale.get('product_life_cycle', 'No Definido')
                    
                    total_venta = abs(float(sale.get('price_subtotal', 0)))
                else:
                    # Odoo: campos vienen como [id, nombre]
                    linea_comercial_original = extraer_rel(sale.get('commercial_line_national_id'))
                    linea_comercial = 'Sin Línea'
                    if linea_comercial_original:
                        # Filtrar VENTA INTERNACIONAL (igual que KPI Venta)
                        if 'VENTA INTERNACIONAL' in linea_comercial_original.upper():
                            continue
                        linea_comercial = normalizar_linea(linea_comercial_original)
                    
                    categoria = extraer_rel(sale.get('categ_id'), 'Sin Categoría')
                    
                    via_administracion = extraer_rel(sale.get('administration_way_id'), 'No Definido')
                    
                    clasificacion = extraer_rel(sale.get('pharmacological_classification_id'), 'No Definido')
                    
                    forma_farmaceutica = extraer_rel(sale.get('pharmaceutical_forms_id'), 'No Definido')
                    
                    linea_produccion = extraer_rel(sale.get('production_line_id'), 'No Definido')
                    
                    ciclo_vida = sale.get('product_life_cycle', 'No Definido')
                    
                    total_venta = abs(float(sale.get('balance', 0)))
                
                if not ciclo_vida or ciclo_vida == '':
                    ciclo_vida = 'No Definido'
                
                if total_venta <= 0:
                    continue
                
                # Agregar a filtros
                filtros_lineas.add(linea_comercial)
                filtros_categorias.add(categoria)
                filtros_ciclos.add(ciclo_vida)
                filtros_vias.add(via_administracion)
                filtros_clasif.add(clasificacion)
                filtros_formas.add(forma_farmaceutica)
                filtros_prod.add(linea_produccion)

                append_registro({
                    'mes_nombre': mes_nombre,
                    'linea_comercial': linea_comercial,
                    'categoria': categoria,
                    'ciclo_vida': ciclo_vida,
                    'via_administracion': via_administracion,
                    'clasificacion': clasificacion,
                    'forma_farmaceutica': forma_farmaceutica,
                    'linea_produccion': linea_produccion,
                    'total': total_venta
                })
                
            except Exception as e:
                continue
        
        print(f"✅ Registros procesados: {len(ventas_detalladas_por_mes)}")
        print(f"📊 Filtros: LC={len(filtros_disponibles['lineas_comerciales'])}, Cat={len(filtros_disponibles['categorias'])}, CV={len(filtros_disponibles['ciclos_vida'])}")
        print(f"🔍 DEBUG: Saliendo del try, preparando return...")
        
    except Exception as e:
        print(f"❌ Error al generar datos de ventas por mes: {e}")
        import traceback
        traceback.print_exc()
    
    print(f"🔍 DEBUG: Después del except, creando diccionario resultado...")
    try:
        resultado = {
            'registros': ventas_detalladas_por_mes,
            'filtros': {
                'lineas_comerciales': sorted([x for x in filtros_disponibles['lineas_comerciales'] if x is not None]),
                'categorias': sorted([x for x in filtros_disponibles['categorias'] if x is not None]),
                'ciclos_vida': sorted([x for x in filtros_disponibles['ciclos_vida'] if x is not None]),
                'vias_administracion': sorted([x for x in filtros_disponibles['vias_administracion'] if x is not None]),
                'clasificaciones': sorted([x for x in filtros_disponibles['clasificaciones'] if x is not None]),
                'formas_farmaceuticas': sorted([x for x in filtros_disponibles['formas_farmaceuticas'] if x is not None]),
                'lineas_produccion': sorted([x for x in filtros_disponibles['lineas_produccion'] if x is not None])
            }
        }
        print(f"🎯 RETORNANDO DATOS VENTAS MES: {len(resultado['registros'])} registros, {len(resultado['filtros']['lineas_comerciales'])} líneas comerciales")
        return resultado
    except Exception as e:
        print(f"❌ ERROR CRÍTICO al crear diccionario resultado: {e}")
        import traceback
        traceback.print_exc()
        return {'registros': [], 'filtros': {}}


@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    try:
        # --- Lógica de Permisos de Administrador ---
        is_admin = session.get('username') in ADMIN_USERS
        
        # --- Parámetro de Período RFM ---
        periodo_rfm = request.args.get('periodo_rfm', '0')  # 0=YTD, 30/90/180/365=días
        try:
            periodo_rfm_dias = int(periodo_rfm)
        except:
            periodo_rfm_dias = 0
        
        print(f"📅 Período RFM seleccionado: {periodo_rfm_dias} días ({'YTD del mes' if periodo_rfm_dias == 0 else f'Últimos {periodo_rfm_dias} días'})")

        # Obtener año actual y año seleccionado
        fecha_actual = datetime.now()
        año_actual = fecha_actual.year
        año_seleccionado = int(request.args.get('año', año_actual))
        
        # Generar lista de años disponibles (desde 2020 hasta el año actual)
        años_disponibles = list(range(2020, año_actual + 1))
        
        # Obtener mes seleccionado (puede venir como "2025-02" o como "2")
        # Si no se especifica mes, usar el mes actual solo si año==año_actual, sino usar enero
        if 'mes' not in request.args:
            if año_seleccionado == año_actual:
                # En Render (Free), abrir por defecto en el último mes CERRADO (de Supabase,
                # ligero) para evitar el OOM/timeout de cargar el mes en curso desde Odoo.
                # El usuario puede seleccionar el mes en curso manualmente.
                if IS_RENDER and fecha_actual.month > 1:
                    mes_default = f"{año_actual}-{(fecha_actual.month - 1):02d}"
                else:
                    mes_default = fecha_actual.strftime('%Y-%m')
            else:
                mes_default = f"{año_seleccionado}-01"
        else:
            mes_default = request.args.get('mes')
        
        mes_param = mes_default
        
        # Si el mes ya tiene formato YYYY-MM, usarlo directamente
        if '-' in str(mes_param) and len(str(mes_param).split('-')) == 2:
            mes_seleccionado = str(mes_param)
        else:
            # Si solo viene el número del mes, construir el formato completo
            mes_seleccionado = f"{año_seleccionado}-{str(mes_param).zfill(2)}"
        
        año_sel, mes_sel = mes_seleccionado.split('-')
        año_sel_int = int(año_sel)
        mes_sel_int = int(mes_sel)
        
        # Ajustar mes si no pertenece al año seleccionado
        if año_sel_int != año_seleccionado:
            mes_seleccionado = f"{año_seleccionado}-01"
            año_sel_int = año_seleccionado
            mes_sel_int = 1
        
        # Determinar fuente de datos (mover antes del check de caché)
        # Pasamos el mes para enrutar meses cerrados de 2026+ a Supabase
        data_source = get_data_source(año_sel_int, mes_sel_int)

        # --- REVISAR CACHÉ ANTES DE HACER CÁLCULOS ---
        # Permitir bypass del caché con parámetro ?nocache=1
        nocache = request.args.get('nocache', '0') == '1'
        if nocache:
            print("🚫 Parámetro nocache=1 detectado, se omitirá el caché")
            cached_data = None
        else:
            cached_data = get_cached_data(año_sel_int, mes_sel_int)
            
        if cached_data:
            cached_data['is_admin'] = is_admin # Re-inyectar datos de sesión
            cached_data['desde_cache'] = True
            cached_data['años_disponibles'] = años_disponibles
            cached_data['año_seleccionado'] = año_seleccionado
            cached_data['mes_seleccionado'] = mes_seleccionado
            
            # Recalcular mes_nombre con los valores correctos
            meses_disponibles_temp = get_meses_del_año(año_seleccionado)
            mes_obj_temp = next((m for m in meses_disponibles_temp if m['key'] == mes_seleccionado), None)
            cached_data['mes_nombre'] = mes_obj_temp['nombre'] if mes_obj_temp else "Mes Desconocido"
            cached_data['meses_disponibles'] = meses_disponibles_temp
            
            # CRÍTICO: Recalcular tendencia_12_meses si el año solicitado es diferente
            tendencia_año_cache = None
            if 'tendencia_12_meses' in cached_data and len(cached_data['tendencia_12_meses']) > 0:
                # Extraer año de la tendencia cacheada (formato: "YYYY-MM")
                primer_mes = cached_data['tendencia_12_meses'][0].get('mes', '')
                if '-' in primer_mes:
                    tendencia_año_cache = int(primer_mes.split('-')[0])
            
            if tendencia_año_cache and tendencia_año_cache != año_seleccionado:
                print(f"🔄 Recalculando tendencia_12_meses: caché tiene año {tendencia_año_cache}, solicitado {año_seleccionado}")
                
                # Recalcular tendencia para el año correcto
                tendencia_12_meses_recalculada = []
                fecha_inicio_tendencia = f"{año_seleccionado}-01-01"
                
                # Tendencia anual: probamos enero para usar Supabase si el año está migrado
                # (el resumen mensual de Odoo está roto: "Función agregada inválida 'month'")
                tendencia_data_source = get_data_source(año_seleccionado, 1)

                # Optimización: Si es año actual con Odoo, solo hasta hoy
                if tendencia_data_source == 'odoo' and año_seleccionado == datetime.now().year:
                    fecha_fin_tendencia = datetime.now().strftime('%Y-%m-%d')
                else:
                    fecha_fin_tendencia = f"{año_seleccionado}-12-31"
                if tendencia_data_source == 'supabase':
                    resumen_mensual = supabase_manager.get_sales_by_month(fecha_inicio_tendencia, fecha_fin_tendencia)
                else:
                    resumen_mensual = data_manager.get_sales_summary_by_month(fecha_inicio_tendencia, fecha_fin_tendencia)
                
                # Obtener metas del año desde la fuente correcta (Google Sheets o Supabase)
                metas_historicas = load_metas(año_seleccionado)
                
                meses_es = {
                    1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio',
                    7: 'julio', 8: 'agosto', 9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
                }
                
                for mes_num in range(1, 13):
                    fecha_mes = datetime(año_seleccionado, mes_num, 1)
                    mes_key = f"{año_seleccionado}-{mes_num:02d}"
                    venta_mes = 0
                    label_busqueda_es = f"{meses_es[mes_num]} {año_seleccionado}"
                    label_busqueda_en = fecha_mes.strftime('%B %Y').lower()
                    
                    for key, val in resumen_mensual.items():
                        key_lower = key.lower()
                        if label_busqueda_es == key_lower or label_busqueda_en == key_lower:
                            venta_mes = val
                            break
                        if meses_es[mes_num] in key_lower and str(año_seleccionado) in key_lower:
                            venta_mes = val
                            break
                    
                    try:
                        meta_key = f"{año_seleccionado}-{mes_num:02d}"
                        metas_mes_data = metas_historicas.get(meta_key, {}).get('metas', {})
                        meta_mes = sum(metas_mes_data.values())
                    except:
                        meta_mes = 0
                    
                    tendencia_12_meses_recalculada.append({
                        'mes': mes_key,
                        'mes_nombre': fecha_mes.strftime('%b %Y'),
                        'venta': venta_mes,
                        'meta': meta_mes,
                        'cumplimiento': (venta_mes / meta_mes * 100) if meta_mes > 0 else 0
                    })
                
                cached_data['tendencia_12_meses'] = tendencia_12_meses_recalculada
                print(f"✅ Tendencia recalculada: {len(tendencia_12_meses_recalculada)} meses para año {año_seleccionado}")
            
            print(f"\n{'='*80}")
            print(f"🎯 RENDER DESDE CACHÉ")
            print(f"   año_seleccionado: {cached_data['año_seleccionado']}")
            print(f"   mes_seleccionado: {cached_data['mes_seleccionado']}")
            print(f"   mes_nombre: {cached_data['mes_nombre']}")
            print(f"   tendencia_12_meses: {len(cached_data.get('tendencia_12_meses', []))} meses")
            if len(cached_data.get('tendencia_12_meses', [])) > 0:
                print(f"   Primer mes tendencia: {cached_data['tendencia_12_meses'][0].get('mes', 'N/A')}")
            print(f"{'='*80}\n")
            
            # Obtener grupos de venta si no están en caché (para Odoo)
            if 'grupos_venta' not in cached_data or not cached_data['grupos_venta']:
                grupos_venta = []
                try:
                    if data_source == 'odoo':
                        grupos_venta = data_manager.models.execute_kw(
                            data_manager.db, data_manager.uid, data_manager.password,
                            'agr.groups', 'search_read',
                            [[]],
                            {'fields': ['id', 'name'], 'order': 'name'}
                        )
                        print(f"📋 Obtenidos {len(grupos_venta)} grupos de venta para filtros (desde caché)")
                except Exception as e:
                    print(f"⚠️ Error obteniendo grupos de venta: {e}")
                cached_data['grupos_venta'] = grupos_venta
            
            # Gráfico de productos: omitir en Render Free (red lenta)
            is_render = os.getenv('RENDER', 'false').lower() == 'true'
            
            if is_render:
                print("⏭️  Gráfico de productos omitido en caché (Render Free Tier)")
                cached_data['datos_ventas_mes_filtros'] = {
                    'datos': [],
                    'filtros': {
                        'lineas_comerciales': [],
                        'categorias': [],
                        'ciclos_vida': [],
                        'vias_administracion': [],
                        'clasificaciones': [],
                        'formas_farmaceuticas': [],
                        'lineas_produccion': []
                    }
                }
            else:
                print(f"📊 Regenerando gráfico de productos desde caché ({data_source})...")
                cached_data['datos_ventas_mes_filtros'] = generar_datos_ventas_mes(año_sel_int, data_source, fecha_actual)
            
            return render_template('dashboard_clean.html', **cached_data)

        # Si no hay caché válido, continuar con la obtención de datos
        print(f"🔄 Mes solicitado ({mes_seleccionado}): No se encontró caché válido. Obteniendo datos frescos...")

        # --- NUEVA LÓGICA DE FILTRADO POR DÍA ---
        # Obtener el día final del filtro, si existe
        dia_fin_param = request.args.get('dia_fin')

        # Crear todos los meses del año seleccionado
        meses_disponibles = get_meses_del_año(año_seleccionado)
        
        # Obtener nombre del mes seleccionado
        mes_obj = next((m for m in meses_disponibles if m['key'] == mes_seleccionado), None)
        mes_nombre = mes_obj['nombre'] if mes_obj else "Mes Desconocido"
        
        print(f"📋 Valores iniciales: año_seleccionado={año_seleccionado}, mes_seleccionado={mes_seleccionado}, mes_nombre={mes_nombre}")
        
        año_sel, mes_sel = mes_seleccionado.split('-')
        año_sel_int = int(año_sel)
        mes_sel_int = int(mes_sel)
        
        # Determinar el día a usar para los cálculos y la fecha final
        if dia_fin_param:
            try:
                dia_actual = int(dia_fin_param)
                fecha_fin = f"{año_sel}-{mes_sel}-{str(dia_actual).zfill(2)}"
            except (ValueError, TypeError):
                # Si el parámetro no es un número válido, usar el comportamiento por defecto
                dia_fin_param = None # Resetear para que entre al siguiente bloque
        
        if not dia_fin_param:
            # Comportamiento original si no hay filtro de día
            if mes_seleccionado == fecha_actual.strftime('%Y-%m'):
                # Mes actual: usar día actual
                dia_actual = fecha_actual.day
            else:
                # Mes pasado: usar último día del mes
                ultimo_dia_mes = calendar.monthrange(año_sel_int, mes_sel_int)[1]
                dia_actual = ultimo_dia_mes
            fecha_fin = f"{año_sel}-{mes_sel}-{str(dia_actual).zfill(2)}"

        fecha_inicio = f"{año_sel}-{mes_sel}-01"
        # --- FIN DE LA NUEVA LÓGICA ---
        
        # --- CALCULAR FECHAS PARA RFM SEGÚN PERÍODO ---
        if periodo_rfm_dias == 0:
            # YTD del mes: desde inicio del mes hasta fecha_fin
            fecha_inicio_rfm = fecha_inicio
            fecha_fin_rfm = fecha_fin
            print(f"📅 RFM usando YTD del mes: {fecha_inicio_rfm} a {fecha_fin_rfm}")
        else:
            # Últimos N días desde fecha_fin
            fecha_fin_obj = dt.strptime(fecha_fin, '%Y-%m-%d')
            fecha_inicio_obj = fecha_fin_obj - timedelta(days=periodo_rfm_dias)
            fecha_inicio_rfm = fecha_inicio_obj.strftime('%Y-%m-%d')
            fecha_fin_rfm = fecha_fin
            print(f"📅 RFM usando últimos {periodo_rfm_dias} días: {fecha_inicio_rfm} a {fecha_fin_rfm}")
        
        # data_source ya se definió antes del check de caché (línea ~460)
        if data_source == 'supabase':
            print(f"📊 Obteniendo datos históricos del {año_sel_int} desde Supabase...")
        else:
            print(f"🔄 Obteniendo datos del {año_sel_int} desde Odoo...")

        # Obtener metas del mes seleccionado desde la fuente correcta (Google Sheets o Supabase)
        metas_historicas = load_metas(año_sel_int)
        metas_del_mes_raw = metas_historicas.get(mes_seleccionado, {}).get('metas', {})
        metas_ipn_del_mes_raw = metas_historicas.get(mes_seleccionado, {}).get('metas_ipn', {})
        
        # Obtener grupos de venta desde Odoo para filtros del mapa
        grupos_venta = []
        try:
            if data_source == 'odoo':
                grupos_venta = data_manager.models.execute_kw(
                    data_manager.db, data_manager.uid, data_manager.password,
                    'agr.groups', 'search_read',
                    [[]],
                    {'fields': ['id', 'name'], 'order': 'name'}
                )
                print(f"📋 Obtenidos {len(grupos_venta)} grupos de venta para filtros")
        except Exception as e:
            print(f"⚠️ Error obteniendo grupos de venta: {e}")
        
        # Consolidar metas de GENVET con TERCEROS
        metas_del_mes = {}
        metas_ipn_del_mes = {}
        
        for linea_id, valor in metas_del_mes_raw.items():
            if linea_id == 'genvet':
                # Sumar la meta de genvet a terceros
                metas_del_mes['terceros'] = metas_del_mes.get('terceros', 0) + valor
            else:
                metas_del_mes[linea_id] = valor
        
        for linea_id, valor in metas_ipn_del_mes_raw.items():
            if linea_id == 'genvet':
                # Sumar la meta IPN de genvet a terceros
                metas_ipn_del_mes['terceros'] = metas_ipn_del_mes.get('terceros', 0) + valor
            else:
                metas_ipn_del_mes[linea_id] = valor
        
        # Las líneas comerciales se generan dinámicamente más adelante.
        
        # Obtener datos reales de ventas desde la fuente correspondiente
        is_render = os.getenv('RENDER', 'false').lower() == 'true'
        sales_data_anual_compartida = None
        try:
            # Las fechas de inicio y fin ahora se calculan más arriba
            
            # En desarrollo local (no Render), cargar el año una sola vez y derivar el mes
            if not is_render:
                fecha_inicio_anual = f"{año_sel_int}-01-01"
                if data_source == 'odoo' and año_sel_int == fecha_actual.year:
                    fecha_fin_anual = f"{año_sel_int}-{fecha_actual.month:02d}-{fecha_actual.day:02d}"
                else:
                    fecha_fin_anual = f"{año_sel_int}-12-31"

                if data_source == 'supabase':
                    sales_data_anual_compartida = supabase_manager.get_dashboard_data(fecha_inicio_anual, fecha_fin_anual)
                else:
                    sales_data_anual_compartida = data_manager.get_sales_lines(
                        date_from=fecha_inicio_anual,
                        date_to=fecha_fin_anual,
                        limit=50000
                    )

                sales_data = []
                for sale in sales_data_anual_compartida:
                    fecha_str = (sale.get('invoice_date') or sale.get('date_order', ''))[:10]
                    if not fecha_str:
                        continue
                    if fecha_inicio <= fecha_str <= fecha_fin:
                        sales_data.append(sale)

                print(f"⚡ Datos del mes derivados de carga anual compartida: {len(sales_data)} líneas")
            # En Render, mantener estrategia liviana mensual
            elif data_source == 'supabase':
                sales_data = supabase_manager.get_dashboard_data(fecha_inicio, fecha_fin)
                print(f"📊 Obtenidas {len(sales_data)} líneas de ventas desde Supabase")
            else:
                # Solo traer ventas estrictamente del mes seleccionado, respetando filtros
                sales_data = data_manager.get_sales_lines(
                    date_from=fecha_inicio,
                    date_to=fecha_fin,
                    limit=SALES_LIMIT
                )
                # Filtrar por mes exacto en caso de que Odoo devuelva líneas fuera del rango
                sales_data = [s for s in sales_data if s.get('invoice_date', '').startswith(f'{año_sel}-{mes_sel}')]  # YYYY-MM
                print(f"📊 Obtenidas {len(sales_data)} líneas de ventas desde Odoo")
            
            # Obtener clientes históricos (cartera activa) - clientes que han comprado desde inicio del año hasta el mes seleccionado
            try:
                # Calcular fecha desde inicio del año hasta el final del mes seleccionado
                fecha_inicio_ano = datetime(año_seleccionado, 1, 1).strftime('%Y-%m-%d')
                ultimo_dia_mes_sel = calendar.monthrange(int(año_sel), int(mes_sel))[1]
                fecha_fin_mes_sel = f"{int(año_sel):04d}-{int(mes_sel):02d}-{ultimo_dia_mes_sel:02d}"
                
                # Determinar fuente de datos según el AÑO SELECCIONADO (no el año actual del sistema)
                source = get_data_source(año_seleccionado, mes_sel_int)
                
                # ESTRATEGIA MIXTA: Para años históricos (2025), usar cartera total de Odoo
                # pero activos desde Supabase para comparar contra la base completa
                if source == 'supabase':
                    print(f"👥 CARTERA: Obteniendo base total desde Odoo (todos los clientes)")
                    total_clientes = data_manager.get_total_partners_count()
                    print(f"👥 Base de cartera total: {total_clientes} clientes")
                else:
                    print(f"👥 Obteniendo cartera de clientes desde Odoo (año {año_seleccionado})")
                    # Para año actual, cartera = clientes que han comprado en el año
                    total_clientes = data_manager.get_active_partners_count(
                        date_from=fecha_inicio_ano,
                        date_to=fecha_fin_mes_sel
                    )
                
                print(f"👥 Total de clientes en cartera (año {año_seleccionado}): {total_clientes}")
            except Exception as e:
                print(f"⚠️ Error obteniendo cartera de clientes: {e}")
                total_clientes = 0
            
        except Exception as e:
            print(f"⚠️ Error obteniendo datos de Odoo: {e}")
            sales_data = []
            total_clientes = 0
        
        # Procesar datos de ventas por línea comercial
        datos_lineas = []
        total_venta = 0
        total_vencimiento = 0
        total_venta_pn = 0
        
        # --- CÁLCULO DE TOTALES ---
        # Calcular totales de metas ANTES de filtrar las líneas para la tabla.
        # Esto asegura que ECOMMERCE se incluya en el total general del KPI.
        total_meta = sum(metas_del_mes.values())
        total_meta_pn = sum(metas_ipn_del_mes.values())
        
        # Mapeo de líneas comerciales de Odoo a IDs locales
        mapeo_lineas = {
            'PETMEDICA': 'petmedica',
            'AGROVET': 'agrovet', 
            'PET NUTRISCIENCE': 'pet_nutriscience',
            'AVIVET': 'avivet',
            'OTROS': 'otros',
            'TERCEROS': 'terceros',
            'INTERPET': 'interpet',
        }
        
        # Calcular ventas reales por línea comercial
        ventas_por_linea = {}
        ventas_por_ruta = {}
        ventas_ipn_por_linea = {} # Nueva variable para ventas de productos nuevos
        ventas_por_producto = {}
        ciclo_vida_por_producto = {}
        ventas_por_ciclo_vida = {}
        ventas_por_forma = {}
        clientes_por_linea = {}  # Nueva variable para contar clientes únicos por línea
        
        # Diccionarios para rastrear por canal (DIGITAL/NACIONAL)
        clientes_por_linea_y_canal = {}  # {linea: {'DIGITAL': set(), 'NACIONAL': set()}}
        ventas_por_linea_y_canal = {}  # {linea: {'DIGITAL': 0, 'NACIONAL': 0}}
        
        ventas_sin_linea = 0
        ventas_sin_canal = 0
        ventas_categoria_excluida = 0
        
        # Categorías a excluir (igual que Proyecto A)
        categorias_excluidas = {315, 333, 304, 314, 318, 339}
        normalizar_linea = normalizar_linea_comercial
        limpiar_atrevia = limpiar_nombre_atrevia
        extraer_rel = extraer_nombre_relacional
        
        # DEBUG: Ver estructura de un registro de ventas
        if sales_data and len(sales_data) > 0:
            print(f"🔍 DEBUG: Estructura del primer registro de ventas:")
            ejemplo = sales_data[0]
            print(f"   - Campos disponibles: {list(ejemplo.keys())}")
            print(f"   - commercial_line_national_id: {ejemplo.get('commercial_line_national_id')}")
            print(f"   - linea_comercial: {ejemplo.get('linea_comercial')}")
            print(f"   - balance: {ejemplo.get('balance')}")
        
        # IMPORTANTE: Los datos de Supabase YA VIENEN FILTRADOS desde Odoo
        # Solo aplicar filtros si la fuente es Odoo
        aplicar_filtros = (data_source == 'odoo')
        if not aplicar_filtros:
            print(f"ℹ️ Fuente: Supabase - Los datos ya vienen filtrados, NO se aplicarán filtros adicionales")
        else:
            print(f"ℹ️ Fuente: Odoo - Se aplicarán filtros de categorías e internacional")
        
        for sale in sales_data:
            sale_get = sale.get
            # Excluir categorías específicas (solo para Odoo)
            if aplicar_filtros:
                categ_id = sale_get('categ_id')
                if categ_id and isinstance(categ_id, list) and len(categ_id) > 0:
                    categ_id_num = categ_id[0]
                    if categ_id_num in categorias_excluidas:
                        ventas_categoria_excluida += 1
                        continue
            
            # Excluir VENTA INTERNACIONAL (solo para Odoo)
            linea_comercial = sale_get('commercial_line_national_id')
            nombre_linea_actual = None
            nombre_linea_extraido = extraer_rel(linea_comercial)
            if nombre_linea_extraido:
                nombre_linea_original = nombre_linea_extraido.upper()
                if aplicar_filtros and 'VENTA INTERNACIONAL' in nombre_linea_original:
                    continue
                # Aplicar normalización para agrupar GENVET y MARCA BLANCA como TERCEROS
                nombre_linea_actual = normalizar_linea(nombre_linea_original)
            # NOTA: Si no tiene linea_comercial, nombre_linea_actual queda None
            # y esa venta NO se sumará a ninguna línea (se ignora silenciosamente)
            
            # También filtrar por canal de ventas (solo para Odoo)
            if aplicar_filtros:
                canal_ventas = sale_get('sales_channel_id')
                if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                    nombre_canal = canal_ventas[1].upper()
                    if 'VENTA INTERNACIONAL' in nombre_canal or 'INTERNACIONAL' in nombre_canal:
                        continue
            
            canal_ventas = sale_get('sales_channel_id')
            if not aplicar_filtros or not (canal_ventas and isinstance(canal_ventas, list)):
                # Contar ventas sin canal pero NO excluir
                ventas_sin_canal += 1
            
            # Procesar el balance de la venta (Odoo usa 'balance', Supabase usa 'price_subtotal')
            balance_float = float(sale_get('balance') or sale_get('price_subtotal', 0))
            if balance_float != 0:
                
                # Sumar a ventas totales por línea
                if nombre_linea_actual:
                    ventas_por_linea[nombre_linea_actual] = ventas_por_linea.get(nombre_linea_actual, 0) + balance_float
                    
                    # Contar clientes únicos por línea comercial
                    partner_name = sale_get('partner_name', '').strip()
                    if partner_name:
                        if nombre_linea_actual not in clientes_por_linea:
                            clientes_por_linea[nombre_linea_actual] = set()
                        clientes_por_linea[nombre_linea_actual].add(partner_name)
                        
                        # Clasificar por canal usando sales_channel_id (Odoo) o canal (Supabase)
                        canal_venta = 'NACIONAL'  # Default
                        
                        # Intentar primero con campo 'canal' de Supabase
                        canal_directo = sale_get('canal')
                        if canal_directo:
                            canal_str = str(canal_directo).upper()
                            if 'DIGITAL' in canal_str or 'E-COMMERCE' in canal_str or 'ECOMMERCE' in canal_str:
                                canal_venta = 'DIGITAL'
                        else:
                            # Si no existe 'canal', usar sales_channel_id de Odoo
                            sales_channel = sale_get('sales_channel_id')
                            if sales_channel and isinstance(sales_channel, list) and len(sales_channel) > 1:
                                channel_name = sales_channel[1].upper()
                                if 'DIGITAL' in channel_name or 'E-COMMERCE' in channel_name or 'ECOMMERCE' in channel_name:
                                    canal_venta = 'DIGITAL'
                        
                        # Inicializar estructuras si no existen
                        if nombre_linea_actual not in clientes_por_linea_y_canal:
                            clientes_por_linea_y_canal[nombre_linea_actual] = {'DIGITAL': set(), 'NACIONAL': set()}
                        if nombre_linea_actual not in ventas_por_linea_y_canal:
                            ventas_por_linea_y_canal[nombre_linea_actual] = {'DIGITAL': 0, 'NACIONAL': 0}
                        
                        # Agregar cliente y venta al canal correspondiente
                        clientes_por_linea_y_canal[nombre_linea_actual][canal_venta].add(partner_name)
                        ventas_por_linea_y_canal[nombre_linea_actual][canal_venta] += balance_float
                
                # LÓGICA FINAL: Sumar si la RUTA (route_id) coincide con los valores especificados
                ruta = sale_get('route_id')
                # Se cambia la comparación al ID de la ruta (ruta[0]) para evitar problemas con traducciones.
                if isinstance(ruta, list) and len(ruta) > 0 and ruta[0] in [18, 19]:
                    if nombre_linea_actual:
                        ventas_por_ruta[nombre_linea_actual] = ventas_por_ruta.get(nombre_linea_actual, 0) + balance_float
                
                # Sumar a ventas de productos nuevos (IPN) - Lógica restaurada
                ciclo_vida = sale_get('product_life_cycle')
                if ciclo_vida and ciclo_vida == 'nuevo':
                    if nombre_linea_actual:
                        ventas_ipn_por_linea[nombre_linea_actual] = ventas_ipn_por_linea.get(nombre_linea_actual, 0) + balance_float
                
                # Agrupar por producto para Top 7
                producto_nombre = sale_get('name', '').strip()
                if producto_nombre:
                    # Limpiar nombres de ATREVIA eliminando indicadores de tamaño/presentación
                    producto_nombre_limpio = limpiar_atrevia(producto_nombre)
                    ventas_por_producto[producto_nombre_limpio] = ventas_por_producto.get(producto_nombre_limpio, 0) + balance_float
                    if producto_nombre_limpio not in ciclo_vida_por_producto:
                        ciclo_vida_por_producto[producto_nombre_limpio] = ciclo_vida
                
                # Agrupar por ciclo de vida para el gráfico de dona
                ciclo_vida_grafico = ciclo_vida if ciclo_vida else 'No definido'
                ventas_por_ciclo_vida[ciclo_vida_grafico] = ventas_por_ciclo_vida.get(ciclo_vida_grafico, 0) + balance_float

        # Debug: Mostrar ventas IPN calculadas
        total_ipn = sum(ventas_ipn_por_linea.values())
        if total_ipn > 0:
            print(f"💊 IPN calculado: S/ {total_ipn:,.2f} distribuido en {len([v for v in ventas_ipn_por_linea.values() if v > 0])} líneas")
        else:
            print(f"⚠️ No se encontraron productos con ciclo_vida='nuevo' en las {len(sales_data)} ventas procesadas")
        
        # --- Preparar datos para tabla de clientes por línea comercial CON FILTRO POR CANAL ---
        print(f"🔍 Preparando tabla de clientes por línea con filtro por canal...")
        datos_clientes_por_linea = []
        
        # Variables para totales generales
        total_ventas_general = 0
        total_clientes_general = set()
        total_ventas_digital_general = 0
        total_clientes_digital_general = set()
        total_ventas_nacional_general = 0
        total_clientes_nacional_general = set()
        
        # Obtener todas las líneas únicas de ventas_por_linea
        for nombre_linea in sorted(ventas_por_linea.keys()):
            venta_total = ventas_por_linea.get(nombre_linea, 0)
            
            # Obtener clientes únicos totales
            clientes_unicos = clientes_por_linea.get(nombre_linea, set())
            num_clientes_total = len(clientes_unicos)
            ticket_promedio_total = (venta_total / num_clientes_total) if num_clientes_total > 0 else 0
            
            # Obtener datos por canal
            clientes_digital = clientes_por_linea_y_canal.get(nombre_linea, {}).get('DIGITAL', set())
            clientes_nacional = clientes_por_linea_y_canal.get(nombre_linea, {}).get('NACIONAL', set())
            venta_digital = ventas_por_linea_y_canal.get(nombre_linea, {}).get('DIGITAL', 0)
            venta_nacional = ventas_por_linea_y_canal.get(nombre_linea, {}).get('NACIONAL', 0)
            
            num_clientes_digital = len(clientes_digital)
            num_clientes_nacional = len(clientes_nacional)
            ticket_promedio_digital = (venta_digital / num_clientes_digital) if num_clientes_digital > 0 else 0
            ticket_promedio_nacional = (venta_nacional / num_clientes_nacional) if num_clientes_nacional > 0 else 0
            
            # Agregar fila TODOS (total)
            datos_clientes_por_linea.append({
                'nombre': nombre_linea,
                'venta': venta_total,
                'num_clientes': num_clientes_total,
                'ticket_promedio': ticket_promedio_total,
                'canal': 'TODOS'
            })
            
            # Agregar fila DIGITAL
            datos_clientes_por_linea.append({
                'nombre': nombre_linea,
                'venta': venta_digital,
                'num_clientes': num_clientes_digital,
                'ticket_promedio': ticket_promedio_digital,
                'canal': 'DIGITAL'
            })
            
            # Agregar fila NACIONAL
            datos_clientes_por_linea.append({
                'nombre': nombre_linea,
                'venta': venta_nacional,
                'num_clientes': num_clientes_nacional,
                'ticket_promedio': ticket_promedio_nacional,
                'canal': 'NACIONAL'
            })
            
            # Acumular en totales generales
            total_ventas_general += venta_total
            total_clientes_general.update(clientes_unicos)
            total_ventas_digital_general += venta_digital
            total_clientes_digital_general.update(clientes_digital)
            total_ventas_nacional_general += venta_nacional
            total_clientes_nacional_general.update(clientes_nacional)
        
        # Agregar fila de TOTALES
        num_clientes_total_general = len(total_clientes_general)
        num_clientes_digital_general = len(total_clientes_digital_general)
        num_clientes_nacional_general = len(total_clientes_nacional_general)
        
        datos_clientes_por_linea.append({
            'nombre': '<strong>TOTAL</strong>',
            'venta': total_ventas_general,
            'num_clientes': num_clientes_total_general,
            'ticket_promedio': (total_ventas_general / num_clientes_total_general) if num_clientes_total_general > 0 else 0,
            'canal': 'TODOS',
            'es_total': True
        })
        
        datos_clientes_por_linea.append({
            'nombre': '<strong>TOTAL</strong>',
            'venta': total_ventas_digital_general,
            'num_clientes': num_clientes_digital_general,
            'ticket_promedio': (total_ventas_digital_general / num_clientes_digital_general) if num_clientes_digital_general > 0 else 0,
            'canal': 'DIGITAL',
            'es_total': True
        })
        
        datos_clientes_por_linea.append({
            'nombre': '<strong>TOTAL</strong>',
            'venta': total_ventas_nacional_general,
            'num_clientes': num_clientes_nacional_general,
            'ticket_promedio': (total_ventas_nacional_general / num_clientes_nacional_general) if num_clientes_nacional_general > 0 else 0,
            'canal': 'NACIONAL',
            'es_total': True
        })
        
        print(f"✅ Tabla de clientes por línea: {len(datos_clientes_por_linea)} filas (incluye fila de TOTALES)")

        # --- Calcular cobertura de clientes ---
        # Primero, obtener el canal de cada cliente desde res.partner
        print("🔍 Obteniendo canales de clientes desde res.partner...")
        clientes_con_canal = {}  # {partner_id: nombre_canal}
        
        # Obtener IDs únicos de clientes
        # --- CÁLCULO DE COBERTURA DE CLIENTES ---
        print("🔍 Calculando cobertura de clientes por canal...")
        
        # Determinar fuente de datos según AÑO SELECCIONADO
        source_cobertura = get_data_source(año_seleccionado, mes_sel_int)
        
        if source_cobertura == 'supabase':
            print("📊 Usando Supabase para cálculo de cobertura")
            # 1. Obtener distribución de TODA la cartera histórica por canal (año completo)
            fecha_inicio_ano = datetime(año_seleccionado, 1, 1).strftime('%Y-%m-%d')
            # ultimo_dia_mes_sel y fecha_fin_mes_sel ya están definidos arriba
            cartera_por_canal = supabase_manager.get_active_partners_by_channel(fecha_inicio_ano, fecha_fin_mes_sel)
            
            # 2. Obtener distribución de clientes ACTIVOS en el periodo seleccionado
            activos_por_canal = supabase_manager.get_active_partners_by_channel(fecha_inicio, fecha_fin)
        else:
            print("📊 Usando Odoo para cálculo de cobertura")
            # 1. Obtener distribución de TODA la cartera histórica por canal (año completo)
            fecha_inicio_ano = datetime(año_seleccionado, 1, 1).strftime('%Y-%m-%d')
            # ultimo_dia_mes_sel y fecha_fin_mes_sel ya están definidos arriba
            cartera_por_canal = data_manager.get_active_partners_by_channel(fecha_inicio_ano, fecha_fin_mes_sel)
            
            # 2. Obtener distribución de clientes ACTIVOS en el periodo seleccionado
            activos_por_canal = data_manager.get_active_partners_by_channel(fecha_inicio, fecha_fin)
        
        # 3. Construir tabla de cobertura
        datos_cobertura_canal = []
        todas_las_llaves_canal = sorted(list(set(cartera_por_canal.keys()) | set(activos_por_canal.keys())))
        
        total_cartera_todos = 0
        total_activos_todos = 0
        
        for canal in todas_las_llaves_canal:
            # Excluir canales internacionales
            if 'INTERNACIONAL' in canal.upper():
                continue
                
            cartera = cartera_por_canal.get(canal, 0)
            activos = activos_por_canal.get(canal, 0)
            cobertura = (activos / cartera * 100) if cartera > 0 else 0
            
            datos_cobertura_canal.append({
                'canal': canal,
                'cartera': cartera,
                'activos': activos,
                'cobertura': cobertura
            })
            
            total_cartera_todos += cartera
            total_activos_todos += activos
            
        # 4. Calcular totales generales y métricas globales
        num_clientes_activos = total_activos_todos
        # Note: total_clientes ya fue calculado arriba por get_active_partners_count
        cobertura_clientes = (num_clientes_activos / total_clientes * 100) if total_clientes > 0 else 0
        
        # Agregar fila de totales a la tabla
        datos_cobertura_canal.append({
            'canal': 'TOTAL GENERAL',
            'cartera': total_cartera_todos,
            'activos': total_activos_todos,
            'cobertura': (total_activos_todos / total_cartera_todos * 100) if total_cartera_todos > 0 else 0,
            'es_total': True
        })
        
        print(f"📊 Cobertura global: {num_clientes_activos} activos de {total_clientes} cartera = {cobertura_clientes:.1f}%")
        print(f"📊 Variables para KPIs - total_clientes: {total_clientes}, num_clientes_activos: {num_clientes_activos}, cobertura: {cobertura_clientes:.2f}%")

        # --- CÁLCULO DE COBERTURA POR GRUPOS (PARA TABLA) ---
        # DESHABILITADO TEMPORALMENTE: Toma mucho tiempo y hace muchas consultas a Odoo
        print(f"⚠️ Cálculo de cobertura por grupos deshabilitado (optimización de performance)")
        datos_cobertura_grupos = []

        # --- CÁLCULO DE FRECUENCIA DE COMPRA POR LÍNEA COMERCIAL ---
        # Usa la misma agrupación que "Análisis de Clientes por Línea Comercial"
        # Frecuencia = Total de Pedidos Únicos / Número de Clientes Activos
        
        print(f"📈 Calculando frecuencia de compra por línea comercial...")
        
        # Diccionarios para almacenar pedidos únicos por línea (usando clientes_por_linea ya existente)
        pedidos_unicos_por_linea = {}  # {linea: set(move_ids)}
        pedidos_por_linea_y_canal = {}  # {linea: {'DIGITAL': set(), 'NACIONAL': set()}}
        
        for sale in sales_data:
            # Obtener línea comercial (misma lógica que ventas_por_linea)
            linea_comercial = sale.get('commercial_line_national_id')
            nombre_linea_actual = None
            
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                nombre_linea_original = linea_comercial[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_linea_original:
                    continue
                nombre_linea_actual = normalizar_linea_comercial(nombre_linea_original)
            
            # También filtrar por canal de ventas
            canal_ventas = sale.get('sales_channel_id')
            if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                nombre_canal = canal_ventas[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_canal or 'INTERNACIONAL' in nombre_canal:
                    continue
            
            if not nombre_linea_actual:
                continue
            
            # Inicializar set si no existe
            if nombre_linea_actual not in pedidos_unicos_por_linea:
                pedidos_unicos_por_linea[nombre_linea_actual] = set()
            
            # Agregar pedido único (move_id)
            move_id = sale.get('move_id')
            if move_id:
                if isinstance(move_id, list):
                    move_id = move_id[0]
                pedidos_unicos_por_linea[nombre_linea_actual].add(move_id)
                
                # Clasificar por canal usando sales_channel_id (Odoo) o canal (Supabase)
                canal_venta = 'NACIONAL'  # Default
                
                # Intentar primero con campo 'canal' de Supabase
                canal_directo = sale.get('canal')
                if canal_directo:
                    canal_str = str(canal_directo).upper()
                    if 'DIGITAL' in canal_str or 'E-COMMERCE' in canal_str or 'ECOMMERCE' in canal_str:
                        canal_venta = 'DIGITAL'
                else:
                    # Si no existe 'canal', usar sales_channel_id de Odoo
                    sales_channel = sale.get('sales_channel_id')
                    if sales_channel and isinstance(sales_channel, list) and len(sales_channel) > 1:
                        channel_name = sales_channel[1].upper()
                        if 'DIGITAL' in channel_name or 'E-COMMERCE' in channel_name or 'ECOMMERCE' in channel_name:
                            canal_venta = 'DIGITAL'
                
                # Inicializar estructuras si no existen
                if nombre_linea_actual not in pedidos_por_linea_y_canal:
                    pedidos_por_linea_y_canal[nombre_linea_actual] = {'DIGITAL': set(), 'NACIONAL': set()}
                
                # Agregar pedido al canal correspondiente
                pedidos_por_linea_y_canal[nombre_linea_actual][canal_venta].add(move_id)
        
        # Calcular frecuencia por línea comercial usando clientes_por_linea ya existente
        datos_frecuencia_linea = []
        total_pedidos_general = 0
        total_clientes_general = 0
        
        # Usar las mismas líneas que ya están en clientes_por_linea
        for linea in sorted(clientes_por_linea.keys()):
            # TODOS (total)
            num_clientes_total = len(clientes_por_linea[linea])
            num_pedidos_total = len(pedidos_unicos_por_linea.get(linea, set()))
            frecuencia_total = (num_pedidos_total / num_clientes_total) if num_clientes_total > 0 else 0
            
            # DIGITAL
            clientes_digital = clientes_por_linea_y_canal.get(linea, {}).get('DIGITAL', set())
            pedidos_digital = pedidos_por_linea_y_canal.get(linea, {}).get('DIGITAL', set())
            num_clientes_digital = len(clientes_digital)
            num_pedidos_digital = len(pedidos_digital)
            frecuencia_digital = (num_pedidos_digital / num_clientes_digital) if num_clientes_digital > 0 else 0
            
            # NACIONAL
            clientes_nacional = clientes_por_linea_y_canal.get(linea, {}).get('NACIONAL', set())
            pedidos_nacional = pedidos_por_linea_y_canal.get(linea, {}).get('NACIONAL', set())
            num_clientes_nacional = len(clientes_nacional)
            num_pedidos_nacional = len(pedidos_nacional)
            frecuencia_nacional = (num_pedidos_nacional / num_clientes_nacional) if num_clientes_nacional > 0 else 0
            
            # Fila TODOS
            datos_frecuencia_linea.append({
                'linea': linea,
                'clientes_activos': num_clientes_total,
                'pedidos': num_pedidos_total,
                'frecuencia': frecuencia_total,
                'canal': 'TODOS'
            })
            
            # Fila DIGITAL
            datos_frecuencia_linea.append({
                'linea': linea,
                'clientes_activos': num_clientes_digital,
                'pedidos': num_pedidos_digital,
                'frecuencia': frecuencia_digital,
                'canal': 'DIGITAL'
            })
            
            # Fila NACIONAL
            datos_frecuencia_linea.append({
                'linea': linea,
                'clientes_activos': num_clientes_nacional,
                'pedidos': num_pedidos_nacional,
                'frecuencia': frecuencia_nacional,
                'canal': 'NACIONAL'
            })
            
            total_pedidos_general += num_pedidos_total
            total_clientes_general += num_clientes_total
        
        # Agregar fila de totales (solo TODOS)
        frecuencia_general = (total_pedidos_general / total_clientes_general) if total_clientes_general > 0 else 0
        datos_frecuencia_linea.append({
            'linea': 'TOTAL GENERAL',
            'clientes_activos': total_clientes_general,
            'pedidos': total_pedidos_general,
            'frecuencia': frecuencia_general,
            'es_total': True,
            'canal': 'TODOS'
        })
        
        print(f"📊 Frecuencia de compra: {len([d for d in datos_frecuencia_linea if d.get('canal') == 'TODOS' and not d.get('es_total')])} líneas comerciales procesadas")
        print(f"📊 Frecuencia general: {frecuencia_general:.2f} pedidos/cliente")

        # --- ANÁLISIS RFM (Recency, Frequency, Monetary) CON SEGMENTACIÓN POR CANAL ---
        print(f"📈 Calculando análisis RFM de clientes con filtro por canal...")
        
        from datetime import datetime as dt, timedelta
        
        # Filtrar sales_data para el período RFM seleccionado
        fecha_inicio_rfm_obj = dt.strptime(fecha_inicio_rfm, '%Y-%m-%d')
        fecha_fin_rfm_obj = dt.strptime(fecha_fin_rfm, '%Y-%m-%d')
        
        sales_data_rfm = []
        for sale in sales_data:
            invoice_date = sale.get('invoice_date')
            if invoice_date:
                if isinstance(invoice_date, str):
                    try:
                        fecha_venta = dt.strptime(invoice_date, '%Y-%m-%d')
                    except:
                        continue
                else:
                    fecha_venta = invoice_date
                
                # Filtrar solo ventas dentro del período RFM
                if fecha_inicio_rfm_obj <= fecha_venta <= fecha_fin_rfm_obj:
                    sales_data_rfm.append(sale)
        
        print(f"📊 Datos para RFM: {len(sales_data_rfm)} ventas en el período {fecha_inicio_rfm} a {fecha_fin_rfm} (de {len(sales_data)} totales)")
        
        # Diccionarios para almacenar datos RFM por cliente
        cliente_recency = {}  # Días desde última compra
        cliente_frequency = {}  # Número de pedidos
        cliente_monetary = {}  # Valor total de compras
        cliente_ultima_fecha = {}  # Fecha de última compra
        cliente_canal = {}  # Canal de cada cliente (DIGITAL/NACIONAL)
        cliente_grupo_venta = {}  # Grupo de venta específico (ECOMMERCE, DISTRIBUIDORES, etc.)
        cliente_partner_id = {}  # ID de partner para obtener canal
        
        # Calcular RFM para cada cliente usando datos del período RFM
        fecha_referencia = fecha_fin_rfm_obj  # Usar fecha fin de RFM como referencia
        
        for sale in sales_data_rfm:  # Usar sales_data_rfm en lugar de sales_data
            partner_name = sale.get('partner_name', '').strip()
            if not partner_name:
                continue
            
            # Excluir VENTA INTERNACIONAL
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                if 'VENTA INTERNACIONAL' in linea_comercial[1].upper():
                    continue
            
            # Guardar partner_id para obtener el canal después
            partner_id = sale.get('partner_id')
            if partner_id:
                if isinstance(partner_id, list):
                    partner_id = partner_id[0]
                cliente_partner_id[partner_name] = partner_id
            
            balance = sale.get('balance', 0)
            if isinstance(balance, str):
                balance = float(balance.replace(',', ''))
            
            move_id = sale.get('move_id')
            if isinstance(move_id, list):
                move_id = move_id[0]
            
            # Obtener fecha de factura
            invoice_date = sale.get('invoice_date')
            if invoice_date:
                if isinstance(invoice_date, str):
                    try:
                        fecha_venta = dt.strptime(invoice_date, '%Y-%m-%d')
                    except:
                        continue
                else:
                    fecha_venta = invoice_date
                
                # Actualizar fecha más reciente
                if partner_name not in cliente_ultima_fecha or fecha_venta > cliente_ultima_fecha[partner_name]:
                    cliente_ultima_fecha[partner_name] = fecha_venta
            
            # Frequency: contar pedidos únicos
            if partner_name not in cliente_frequency:
                cliente_frequency[partner_name] = set()
            cliente_frequency[partner_name].add(move_id)
            
            # Monetary: sumar valor
            cliente_monetary[partner_name] = cliente_monetary.get(partner_name, 0) + balance
        
        # Calcular recency (días desde última compra)
        for partner_name, ultima_fecha in cliente_ultima_fecha.items():
            dias = (fecha_referencia - ultima_fecha).days
            cliente_recency[partner_name] = dias
        
        # OBTENER GRUPOS DE VENTA DESDE ODOO PARA TODOS LOS AÑOS (incluye Supabase)
        # Los grupos están en Odoo y son consistentes para todos los períodos
        print(f"🔄 Obteniendo grupos de venta desde Odoo para clasificación...")
        
        try:
            # Lista de partner_ids únicos
            partner_ids_list = list(set(cliente_partner_id.values()))
            
            if partner_ids_list:
                # Consultar res.partner para obtener groups_ids
                partners_info = data_manager.models.execute_kw(
                    data_manager.db, data_manager.uid, data_manager.password,
                    'res.partner', 'search_read',
                    [[('id', 'in', partner_ids_list)]],
                    {'fields': ['id', 'name', 'groups_ids']}
                )
                
                # Consultar agr.groups para obtener nombres
                groups_data = data_manager.models.execute_kw(
                    data_manager.db, data_manager.uid, data_manager.password,
                    'agr.groups', 'search_read',
                    [[]],
                    {'fields': ['id', 'name']}
                )
                
                # Mapear group_id -> nombre
                group_name_map = {g['id']: g['name'] for g in groups_data}
                
                # Mapear partner_id -> canal y grupo
                partner_canal_map = {}
                partner_grupo_map = {}
                canales_encontrados = {'DIGITAL': 0, 'NACIONAL': 0}
                grupos_sin_asignar = 0
                
                for partner in partners_info:
                    groups_ids = partner.get('groups_ids', [])
                    canal = 'NACIONAL'  # Default: NACIONAL
                    grupo_venta = 'SIN GRUPO'
                    
                    if groups_ids:
                        # Tomar el primer grupo
                        first_group_id = groups_ids[0]
                        grupo_venta = group_name_map.get(first_group_id, 'SIN GRUPO')
                        
                        # Clasificar en DIGITAL o NACIONAL según el nombre del grupo
                        if grupo_venta.upper() in ['ECOMMERCE', 'AIRBNB', 'EMPLEADOS']:
                            canal = 'DIGITAL'
                        else:
                            canal = 'NACIONAL'
                    else:
                        grupos_sin_asignar += 1
                    
                    partner_canal_map[partner['id']] = canal
                    partner_grupo_map[partner['id']] = grupo_venta
                    canales_encontrados[canal] = canales_encontrados.get(canal, 0) + 1
                
                # Asignar canal y grupo a cada cliente
                for partner_name, partner_id in cliente_partner_id.items():
                    canal = partner_canal_map.get(partner_id, 'NACIONAL')
                    grupo = partner_grupo_map.get(partner_id, 'SIN GRUPO')
                    cliente_canal[partner_name] = canal
                    cliente_grupo_venta[partner_name] = grupo
                
                # Debug
                grupos_ejemplo = list(set(partner_grupo_map.values()))[:10]
                print(f"📋 Ejemplos de grupos encontrados: {grupos_ejemplo}")
                print(f"⚠️ Clientes SIN GRUPO asignado: {grupos_sin_asignar}")
                print(f"✅ Grupos de venta asignados desde Odoo:")
                print(f"   - DIGITAL: {canales_encontrados['DIGITAL']} clientes")
                print(f"   - NACIONAL: {canales_encontrados['NACIONAL']} clientes")
            else:
                print("⚠️ No se encontraron partner_ids para consultar grupos")
                
        except Exception as e:
            print(f"⚠️ Error obteniendo grupos desde Odoo: {e}")
            # Si falla, usar valores por defecto
            for partner_name in cliente_partner_id.keys():
                if partner_name not in cliente_canal:
                    cliente_canal[partner_name] = 'NACIONAL'
                if partner_name not in cliente_grupo_venta:
                    cliente_grupo_venta[partner_name] = 'SIN GRUPO'

        # Obtener canal de cada cliente según la fuente de datos
        print(f"🔍 Obteniendo canal de {len(cliente_partner_id)} clientes...")
        
        if año_seleccionado >= 2026 and cliente_partner_id:
            # Para Odoo (2026+), obtener grupo de venta desde res.partner.groups_ids
            try:
                partner_ids_list = list(set(cliente_partner_id.values()))
                
                # Consultar partners con sus grupos
                partners_info = data_manager.models.execute_kw(
                    data_manager.db, data_manager.uid, data_manager.password,
                    'res.partner', 'search_read',
                    [[('id', 'in', partner_ids_list)]],
                    {'fields': ['id', 'name', 'groups_ids'], 'context': {'lang': 'es_PE'}}
                )
                
                # Obtener nombres de todos los grupos
                groups_data = data_manager.models.execute_kw(
                    data_manager.db, data_manager.uid, data_manager.password,
                    'agr.groups', 'search_read',
                    [[]],
                    {'fields': ['id', 'name']}
                )
                
                # Mapear group_id -> nombre
                group_name_map = {g['id']: g['name'] for g in groups_data}
                
                # Mapear partner_id -> canal y grupo
                partner_canal_map = {}
                partner_grupo_map = {}  # Mapeo partner_id -> nombre grupo de venta
                canales_encontrados = {'DIGITAL': 0, 'NACIONAL': 0}
                grupos_sin_asignar = 0  # Contador de clientes sin grupo
                
                for partner in partners_info:
                    groups_ids = partner.get('groups_ids', [])
                    canal = 'NACIONAL'  # Default: NACIONAL
                    grupo_venta = 'SIN GRUPO'
                    
                    if groups_ids:
                        # Tomar el primer grupo (un cliente puede tener múltiples grupos)
                        first_group_id = groups_ids[0]
                        grupo_venta = group_name_map.get(first_group_id, 'SIN GRUPO')
                        
                        # Clasificar en DIGITAL o NACIONAL según el nombre del grupo
                        if grupo_venta.upper() in ['ECOMMERCE', 'AIRBNB', 'EMPLEADOS']:
                            canal = 'DIGITAL'
                        else:
                            canal = 'NACIONAL'
                    else:
                        grupos_sin_asignar += 1
                    
                    partner_canal_map[partner['id']] = canal
                    partner_grupo_map[partner['id']] = grupo_venta
                    canales_encontrados[canal] = canales_encontrados.get(canal, 0) + 1
                
                # Asignar canal y grupo a cada cliente
                for partner_name, partner_id in cliente_partner_id.items():
                    canal = partner_canal_map.get(partner_id, 'NACIONAL')
                    grupo = partner_grupo_map.get(partner_id, 'SIN GRUPO')
                    cliente_canal[partner_name] = canal
                    cliente_grupo_venta[partner_name] = grupo
                
                # Debug: Mostrar algunos ejemplos de grupos
                grupos_ejemplo = list(set(partner_grupo_map.values()))[:10]
                print(f"📋 Ejemplos de grupos encontrados: {grupos_ejemplo}")
                print(f"⚠️ Clientes SIN GRUPO asignado en Odoo: {grupos_sin_asignar}")
                
                print(f"✅ Grupos de venta asignados desde Odoo (agr.groups):")
                print(f"   - DIGITAL: {canales_encontrados['DIGITAL']} clientes")
                print(f"   - NACIONAL: {canales_encontrados['NACIONAL']} clientes")
            except Exception as e:
                print(f"⚠️ Error obteniendo canales de Odoo: {e}")
                # Si falla, asignar 'NACIONAL' a los que no tienen
                for partner_name in cliente_partner_id.keys():
                    if partner_name not in cliente_canal:
                        cliente_canal[partner_name] = 'NACIONAL'
                    if partner_name not in cliente_grupo_venta:
                        cliente_grupo_venta[partner_name] = 'SIN GRUPO'
        
        # Separar clientes por canal para calcular percentiles independientes
        clientes_digital = {k: v for k, v in cliente_monetary.items() if cliente_canal.get(k, '').upper() == 'DIGITAL'}
        clientes_nacional = {k: v for k, v in cliente_monetary.items() if cliente_canal.get(k, '').upper() == 'NACIONAL'}
        clientes_otros = {k: v for k, v in cliente_monetary.items() if cliente_canal.get(k, '').upper() not in ['DIGITAL', 'NACIONAL']}
        
        print(f"📊 Distribución: {len(clientes_digital)} DIGITAL, {len(clientes_nacional)} NACIONAL, {len(clientes_otros)} OTROS")
        
        # Calcular factor de ajuste para umbrales según período
        # Umbrales base diseñados para ~30 días (mes completo)
        # Períodos: 0=YTD (~15 días promedio), 30, 90, 180, 365
        if periodo_rfm_dias == 0:
            # YTD del mes: asumir ~15 días promedio
            factor_recency = 0.5  # La mitad del mes
            factor_frequency = 0.5
        else:
            # Escalar proporcionalmente al período de 30 días
            factor_recency = periodo_rfm_dias / 30.0
            factor_frequency = periodo_rfm_dias / 30.0
        
        print(f"📏 Factores de ajuste de umbrales: Recency={factor_recency:.2f}x, Frequency={factor_frequency:.2f}x")
        
        # Crear lista de clientes con sus métricas RFM
        clientes_rfm = []
        for partner_name in cliente_monetary.keys():
            recency = cliente_recency.get(partner_name, 999)
            frequency = len(cliente_frequency.get(partner_name, set()))
            monetary = cliente_monetary.get(partner_name, 0)
            canal = cliente_canal.get(partner_name, 'SIN CANAL')
            
            # Calcular scores RFM (1-3, donde 3 es mejor) CON UMBRALES DIFERENCIADOS POR CANAL Y AJUSTADOS POR PERÍODO
            
            # RECENCY: menor es mejor (días desde última compra) - Ajustado por período
            if canal == 'DIGITAL':
                # Clientes digitales: compras más frecuentes esperadas
                umbral_r3 = int(20 * factor_recency)
                umbral_r2 = int(45 * factor_recency)
                if recency <= umbral_r3:
                    r_score = 3
                elif recency <= umbral_r2:
                    r_score = 2
                else:
                    r_score = 1
            elif canal == 'NACIONAL':
                # Distribuidores: ciclos de compra más largos
                umbral_r3 = int(60 * factor_recency)
                umbral_r2 = int(120 * factor_recency)
                if recency <= umbral_r3:
                    r_score = 3
                elif recency <= umbral_r2:
                    r_score = 2
                else:
                    r_score = 1
            else:
                # Canal no identificado: umbrales intermedios
                umbral_r3 = int(30 * factor_recency)
                umbral_r2 = int(60 * factor_recency)
                if recency <= umbral_r3:
                    r_score = 3
                elif recency <= umbral_r2:
                    r_score = 2
                else:
                    r_score = 1
            
            # FREQUENCY: mayor es mejor (número de pedidos en el período) - Ajustado por período
            if canal == 'DIGITAL':
                # Clientes digitales: se espera mayor frecuencia
                umbral_f3 = max(1, int(4 * factor_frequency))
                umbral_f2 = max(1, int(2 * factor_frequency))
                if frequency >= umbral_f3:
                    f_score = 3
                elif frequency >= umbral_f2:
                    f_score = 2
                else:
                    f_score = 1
            elif canal == 'NACIONAL':
                # Distribuidores: menor frecuencia pero pedidos grandes
                umbral_f3 = max(1, int(2 * factor_frequency))
                umbral_f2 = 1
                if frequency >= umbral_f3:
                    f_score = 3
                elif frequency >= umbral_f2:
                    f_score = 2
                else:
                    f_score = 1
            else:
                # Canal no identificado: umbrales intermedios
                umbral_f3 = max(1, int(3 * factor_frequency))
                umbral_f2 = max(1, int(2 * factor_frequency))
                if frequency >= umbral_f3:
                    f_score = 3
                elif frequency >= umbral_f2:
                    f_score = 2
                else:
                    f_score = 1
            
            # MONETARY: mayor es mejor (calcular por canal para comparación justa)
            if canal == 'DIGITAL' and clientes_digital:
                valores_canal = sorted([v for v in clientes_digital.values()], reverse=True)
            elif canal == 'NACIONAL' and clientes_nacional:
                valores_canal = sorted([v for v in clientes_nacional.values()], reverse=True)
            else:
                valores_canal = sorted([v for v in cliente_monetary.values()], reverse=True)
            
            percentil_33 = valores_canal[len(valores_canal) // 3] if len(valores_canal) >= 3 else 0
            percentil_66 = valores_canal[len(valores_canal) * 2 // 3] if len(valores_canal) >= 3 else 0
            
            if monetary >= percentil_33:
                m_score = 3
            elif monetary >= percentil_66:
                m_score = 2
            else:
                m_score = 1
            
            # Segmentar clientes según RFM
            rfm_segment = f"{r_score}{f_score}{m_score}"
            
            # Mapeo de segmentos a categorías
            if rfm_segment in ['333', '332', '323', '233']:
                categoria = 'Campeones'
                color = '#52c41a'
            elif rfm_segment in ['331', '322', '313', '232', '223']:
                categoria = 'Leales'
                color = '#73d13d'
            elif rfm_segment in ['321', '312', '311', '221', '213', '212']:
                categoria = 'Potenciales'
                color = '#95de64'
            elif rfm_segment in ['231', '222', '211']:
                categoria = 'Nuevos'
                color = '#1890ff'
            elif rfm_segment in ['133', '132', '131', '123']:
                categoria = 'En Riesgo'
                color = '#faad14'
            elif rfm_segment in ['122', '113', '112', '121']:
                categoria = 'Hibernando'
                color = '#ff7a45'
            elif rfm_segment in ['111']:
                categoria = 'Perdidos'
                color = '#ff4d4f'
            else:
                categoria = 'Otros'
                color = '#bfbfbf'
            
            clientes_rfm.append({
                'cliente': partner_name,
                'canal': canal,
                'grupo': cliente_grupo_venta.get(partner_name, 'SIN GRUPO'),
                'recency': recency,
                'frequency': frequency,
                'monetary': monetary,
                'r_score': r_score,
                'f_score': f_score,
                'm_score': m_score,
                'segmento': rfm_segment,
                'categoria': categoria,
                'color': color
            })
        
        # Ordenar por valor monetario
        clientes_rfm_sorted = sorted(clientes_rfm, key=lambda x: x['monetary'], reverse=True)
        
        # Estadísticas de segmentación
        segmentos_rfm = {}
        for cliente in clientes_rfm:
            cat = cliente['categoria']
            if cat not in segmentos_rfm:
                segmentos_rfm[cat] = {'count': 0, 'valor': 0, 'color': cliente['color']}
            segmentos_rfm[cat]['count'] += 1
            segmentos_rfm[cat]['valor'] += cliente['monetary']
        
        # Estadísticas por canal
        segmentos_por_canal = {
            'DIGITAL': {},
            'NACIONAL': {},
            'OTROS': {},
            'TODOS': segmentos_rfm
        }
        
        for cliente in clientes_rfm:
            canal_tipo = cliente['canal']
            if canal_tipo not in ['DIGITAL', 'NACIONAL']:
                canal_tipo = 'OTROS'
            
            cat = cliente['categoria']
            if cat not in segmentos_por_canal[canal_tipo]:
                segmentos_por_canal[canal_tipo][cat] = {'count': 0, 'valor': 0, 'color': cliente['color']}
            segmentos_por_canal[canal_tipo][cat]['count'] += 1
            segmentos_por_canal[canal_tipo][cat]['valor'] += cliente['monetary']
        
        print(f"📊 Análisis RFM: {len(clientes_rfm)} clientes segmentados en {len(segmentos_rfm)} categorías")
        print(f"   - DIGITAL: {len(clientes_digital)} clientes")
        print(f"   - NACIONAL: {len(clientes_nacional)} clientes")
        print(f"   - OTROS: {len(clientes_otros)} clientes")
        

        # --- TENDENCIA HISTÓRICA (12 MESES DEL AÑO SELECCIONADO) ---
        print(f"📈 Generando tendencia histórica de ventas para el año {año_seleccionado}...")
        tendencia_12_meses = []
        fecha_inicio_tendencia = f"{año_seleccionado}-01-01"
        
        # Obtener resumen solo del año seleccionado (no últimos 12 meses mezclados)
        # Tendencia anual: probamos enero para usar Supabase si el año está migrado
        # (el resumen mensual de Odoo está roto: "Función agregada inválida 'month'")
        tendencia_data_source = get_data_source(año_seleccionado, 1)
        
        # Optimización: Si es año actual con Odoo, solo hasta hoy
        if tendencia_data_source == 'odoo' and año_seleccionado == fecha_actual.year:
            fecha_fin_tendencia = fecha_actual.strftime('%Y-%m-%d')
            print(f"📅 Tendencia: Año actual, consultando solo hasta {fecha_fin_tendencia}")
        else:
            fecha_fin_tendencia = f"{año_seleccionado}-12-31"
        if tendencia_data_source == 'supabase':
            # Para Supabase, usar la misma función que el KPI de ventas (ya tiene filtros correctos)
            resumen_mensual = supabase_manager.get_sales_by_month(fecha_inicio_tendencia, fecha_fin_tendencia)
        else:
            resumen_mensual = data_manager.get_sales_summary_by_month(fecha_inicio_tendencia, fecha_fin_tendencia)
        
        for mes_num in range(1, 13):
            fecha_mes = datetime(año_seleccionado, mes_num, 1)
            mes_key = f"{año_seleccionado}-{mes_num:02d}"
            meses_es = {
                1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio',
                7: 'julio', 8: 'agosto', 9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
            }
            
            # Buscar la venta del mes en el resumen
            venta_mes = 0
            # Formato esperado: "enero 2025", "febrero 2025", etc.
            label_busqueda = f"{meses_es[mes_num]} {año_seleccionado}"
            venta_mes = resumen_mensual.get(label_busqueda, 0)
            
            # DEBUG: Ver qué está buscando y qué encuentra
            if mes_num == 1:  # Solo enero para no llenar logs
                print(f"🔍 TENDENCIA DEBUG - Buscando: '{label_busqueda}'")
                print(f"🔍 TENDENCIA DEBUG - Encontrado: {venta_mes}")
                print(f"🔍 TENDENCIA DEBUG - Keys disponibles en resumen_mensual: {list(resumen_mensual.keys())[:3]}")
            
            try:
                meta_key = f"{año_seleccionado}-{mes_num:02d}"
                metas_mes_data = metas_historicas.get(meta_key, {}).get('metas', {})
                meta_mes = sum(metas_mes_data.values())
            except:
                meta_mes = 0
            tendencia_12_meses.append({
                'mes': mes_key,
                'mes_nombre': fecha_mes.strftime('%b %Y'),
                'venta': venta_mes,
                'meta': meta_mes,
                'cumplimiento': (venta_mes / meta_mes * 100) if meta_mes > 0 else 0
            })
        print(f"📊 Tendencia histórica: {len(tendencia_12_meses)} meses procesados para el año {año_seleccionado}")
        
        # --- HEATMAP DE ACTIVIDAD DE VENTAS ---
        print(f"🔥 Generando heatmap de actividad de ventas para {mes_seleccionado}...")
        print(f"   📅 Rango de fechas para heatmap: {fecha_inicio} hasta {fecha_fin}")
        print(f"   📊 Total líneas a procesar: {len(sales_data)}")
        
        # Calcular el primer lunes del mes para alinear semanas con calendario
        primer_dia_mes = dt(año_sel_int, mes_sel_int, 1)
        dia_semana_primer_dia = primer_dia_mes.weekday()  # 0=Lun, 6=Dom
        
        # Calcular qué día del mes es el primer lunes
        if dia_semana_primer_dia == 0:  # El día 1 es lunes
            primer_lunes = 1
        else:
            # Días hasta el próximo lunes
            dias_hasta_lunes = (7 - dia_semana_primer_dia) % 7
            if dias_hasta_lunes == 0:
                dias_hasta_lunes = 7
            primer_lunes = 1 + dias_hasta_lunes
        
        print(f"   📅 Primer lunes del mes: día {primer_lunes}")
        
        # Matriz: Día de semana (0=Lun, 6=Dom) x Semana del mes (0-4)
        heatmap_data = [[0 for _ in range(7)] for _ in range(5)]  # 5 semanas x 7 días
        heatmap_count = [[0 for _ in range(7)] for _ in range(5)]  # Contador para promedios
        
        # Obtener equipos de ventas y sus miembros
        print(f"👥 Obteniendo equipos de ventas y sus miembros...")
        equipos_ventas = {}
        vendedor_a_equipo = {}  # {vendedor_id: equipo_id}
        vendedor_nombre_a_equipo = {}  # {nombre_vendedor: equipo_id} para datos de Supabase
        
        # Crear equipo especial para vendedores sin equipo
        EQUIPO_SIN_ASIGNAR = 'sin_equipo'
        equipos_ventas[EQUIPO_SIN_ASIGNAR] = {
            'nombre': 'SIN EQUIPO',
            'nombre_original': 'SIN EQUIPO',
            'miembros': [],
            'total_ventas': 0
        }
        
        try:
            # Obtener todos los equipos de ventas (incluyendo el líder del equipo)
            teams_ids = data_manager.models.execute_kw(
                data_manager.db, data_manager.uid, data_manager.password,
                'crm.team', 'search_read',
                [[]],
                {'fields': ['id', 'name', 'member_ids', 'user_id']}
            )
            
            # Mapeo de traducción para nombres que puedan venir en inglés
            traducciones_equipos = {
                'Sales': 'AGROVET',
                'INTERPET': 'INTERPET',
                'PETMEDICA': 'PETMEDICA',
                'ECOMMERCE': 'ECOMMERCE',
                'PETNUTRISCIENCE': 'PETNUTRISCIENCE',
                'AIRBNB': 'AIRBNB',
                'ASUNTOS REGULATORIOS': 'ASUNTOS REGULATORIOS',
                'MARCA BLANCA': 'MARCA BLANCA',
                'MARKETING': 'MARKETING',
                'VENTA INTERNACIONAL': 'VENTA INTERNACIONAL',
                'OFICINA': 'OFICINA',
                'AVIVET': 'AVIVET'
            }
            
            for team in teams_ids:
                team_id = team['id']
                team_name_original = team['name']
                # Aplicar traducción si existe
                team_name = traducciones_equipos.get(team_name_original, team_name_original)
                member_ids = team.get('member_ids', [])
                
                # Obtener el líder del equipo (user_id)
                team_leader = team.get('user_id')
                leader_id = None
                if team_leader and isinstance(team_leader, list) and len(team_leader) > 0:
                    leader_id = team_leader[0]
                
                # Crear lista de todos los miembros (incluyendo al líder)
                all_members = list(member_ids)
                if leader_id and leader_id not in all_members:
                    all_members.append(leader_id)
                
                equipos_ventas[team_id] = {
                    'nombre': team_name,
                    'nombre_original': team_name_original,
                    'miembros': all_members,  # Ahora incluye al líder
                    'total_ventas': 0
                }
                
                # Mapear vendedores a equipos por ID (miembros + líder)
                for member_id in all_members:
                    vendedor_a_equipo[str(member_id)] = team_id
                
            # Obtener nombres de los usuarios (vendedores) para mapeo por nombre
            if vendedor_a_equipo:
                user_ids = list(set([int(vid) for vid in vendedor_a_equipo.keys()]))
                users = data_manager.models.execute_kw(
                    data_manager.db, data_manager.uid, data_manager.password,
                    'res.users', 'read',
                    [user_ids],
                    {'fields': ['id', 'name']}
                )
                
                # Crear mapeo nombre → equipo
                for user in users:
                    user_id = str(user['id'])
                    user_name = user['name']
                    if user_id in vendedor_a_equipo:
                        vendedor_nombre_a_equipo[user_name] = vendedor_a_equipo[user_id]
                        print(f"   🔗 Mapeado: {user_name} → {equipos_ventas[vendedor_a_equipo[user_id]]['nombre']}")
                
            print(f"👥 Equipos de ventas encontrados: {len(equipos_ventas) - 1}")  # -1 para excluir SIN EQUIPO
            for tid, tdata in list(equipos_ventas.items())[:5]:
                if tid != EQUIPO_SIN_ASIGNAR:
                    print(f"   - {tdata['nombre']}: {len(tdata['miembros'])} miembros (IDs: {tdata['miembros']})")
        except Exception as e:
            print(f"⚠️ Error obteniendo equipos de ventas: {e}")
        
        # Rastrear vendedores y sus ventas
        vendedores_heatmap = {}  # {vendedor_id: {nombre, total_ventas, equipo_id}}
        heatmap_por_vendedor = {}  # {vendedor_id: [[ventas por día/semana]]}
        heatmap_count_por_vendedor = {}  # {vendedor_id: [[transacciones por día/semana]]}
        ventas_sin_vendedor = 0
        
        transacciones_procesadas = 0
        ventas_excluidas_internacional = 0
        
        for sale in sales_data:
            invoice_date = sale.get('invoice_date')
            if not invoice_date:
                continue
            
            if isinstance(invoice_date, str):
                try:
                    fecha_venta = dt.strptime(invoice_date, '%Y-%m-%d')
                except:
                    continue
            else:
                fecha_venta = invoice_date
            
            balance = sale.get('balance') or sale.get('price_subtotal', 0)
            if isinstance(balance, str):
                balance = float(balance.replace(',', ''))
            
            # Obtener información del vendedor
            user_info = sale.get('invoice_user_id')
            vendedor_id = None
            vendedor_nombre = 'Sin asignar'
            if user_info and isinstance(user_info, list) and len(user_info) >= 2:
                vendedor_id_raw = str(user_info[0])
                vendedor_nombre = user_info[1]
                
                # Si el ID es "0" (datos de Supabase), usar el nombre como ID único
                if vendedor_id_raw == "0":
                    vendedor_id = f"supabase_{vendedor_nombre.replace(' ', '_')}"
                else:
                    vendedor_id = vendedor_id_raw
                
                # Registrar vendedor
                if vendedor_id not in vendedores_heatmap:
                    # Obtener equipo del vendedor
                    equipo_id = None
                    
                    # Si es de Odoo (ID numérico), buscar por ID
                    if vendedor_id_raw != "0":
                        equipo_id = vendedor_a_equipo.get(vendedor_id_raw, None)
                    else:
                        # Si es de Supabase (ID = 0), buscar por nombre
                        equipo_id = vendedor_nombre_a_equipo.get(vendedor_nombre, None)
                    
                    # Si no tiene equipo, asignar al equipo "SIN EQUIPO"
                    if not equipo_id:
                        equipo_id = EQUIPO_SIN_ASIGNAR
                    
                    equipo_nombre = equipos_ventas.get(equipo_id, {}).get('nombre', 'Sin equipo')
                    
                    vendedores_heatmap[vendedor_id] = {
                        'nombre': vendedor_nombre,
                        'total_ventas': 0,
                        'equipo_id': equipo_id,
                        'equipo_nombre': equipo_nombre
                    }
                    heatmap_por_vendedor[vendedor_id] = [[0 for _ in range(7)] for _ in range(5)]
                    heatmap_count_por_vendedor[vendedor_id] = [[0 for _ in range(7)] for _ in range(5)]
                    print(f"   📌 Nuevo vendedor detectado: {vendedor_nombre} (ID: {vendedor_id}) - Equipo: {equipo_nombre}")
                
                vendedores_heatmap[vendedor_id]['total_ventas'] += balance
                
                # Actualizar ventas del equipo (ahora siempre hay un equipo, incluso si es "SIN EQUIPO")
                equipo_id = vendedores_heatmap[vendedor_id].get('equipo_id')
                if equipo_id and equipo_id in equipos_ventas:
                    equipos_ventas[equipo_id]['total_ventas'] += balance
            else:
                ventas_sin_vendedor += 1
            
            # Día de la semana (0=Lunes, 6=Domingo)
            dia_semana = fecha_venta.weekday()
            
            # Calcular semana del mes basándose en calendario real
            dia_mes = fecha_venta.day
            if dia_mes < primer_lunes:
                # Días antes del primer lunes = Semana 1
                semana_mes = 0
            else:
                # Días desde el primer lunes, cada 7 días = nueva semana
                dias_desde_primer_lunes = dia_mes - primer_lunes
                semana_mes = 1 + (dias_desde_primer_lunes // 7)
                semana_mes = min(semana_mes, 4)  # Máximo 5 semanas (índice 0-4)
            
            heatmap_data[semana_mes][dia_semana] += balance
            heatmap_count[semana_mes][dia_semana] += 1
            
            # Agregar a matriz del vendedor
            if vendedor_id:
                heatmap_por_vendedor[vendedor_id][semana_mes][dia_semana] += balance
                heatmap_count_por_vendedor[vendedor_id][semana_mes][dia_semana] += 1
            
            transacciones_procesadas += 1
        
        # Preparar datos para el frontend (formato para ECharts heatmap)
        heatmap_ventas = []
        dias_labels = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
        semanas_labels = ['Semana 1', 'Semana 2', 'Semana 3', 'Semana 4', 'Semana 5']
        
        max_venta_dia = 0
        celdas_activas = 0
        total_ventas_mes = 0
        ultimo_dia_mes = calendar.monthrange(año_sel_int, mes_sel_int)[1]
        
        # Crear un mapa de (semana_idx, dia_idx) -> dia_del_mes
        # Esto nos permite calcular el día del mes para cada celda del heatmap
        mapa_dias = {}
        
        # Para cada día del mes, calcular en qué celda del heatmap debe aparecer
        for dia_del_mes in range(1, ultimo_dia_mes + 1):
            fecha = dt(año_sel_int, mes_sel_int, dia_del_mes)
            dia_semana = fecha.weekday()  # 0=Lun, 6=Dom
            
            # Calcular semana usando la misma lógica que al procesar ventas
            if dia_del_mes < primer_lunes:
                semana_mes = 0
            else:
                dias_desde_primer_lunes = dia_del_mes - primer_lunes
                semana_mes = 1 + (dias_desde_primer_lunes // 7)
                semana_mes = min(semana_mes, 4)
            
            mapa_dias[(semana_mes, dia_semana)] = dia_del_mes
        
        # Debug: mostrar mapa de días para semana 0
        print(f"   🗓️ Mapa de días - Semana 0: {[(k, v) for k, v in mapa_dias.items() if k[0] == 0]}")
        
        # Ahora construir el array para el frontend
        for semana_idx in range(5):
            for dia_idx in range(7):
                # Buscar si existe un día del mes para esta celda
                dia_del_mes = mapa_dias.get((semana_idx, dia_idx))
                
                venta_total = heatmap_data[semana_idx][dia_idx]
                count = heatmap_count[semana_idx][dia_idx]
                
                if dia_del_mes:
                    # Celda válida con un día real del mes
                    if count > 0:
                        celdas_activas += 1
                    
                    total_ventas_mes += venta_total
                    
                    heatmap_ventas.append({
                        'semana': semana_idx,
                        'dia': dia_idx,
                        'valor': venta_total,
                        'transacciones': count,
                        'dia_mes': dia_del_mes
                    })
                    
                    if venta_total > max_venta_dia:
                        max_venta_dia = venta_total
                else:
                    # Celda vacía (día que no existe en este mes)
                    heatmap_ventas.append({
                        'semana': semana_idx,
                        'dia': dia_idx,
                        'valor': -1,  # -1 indica celda no válida
                        'transacciones': 0,
                        'dia_mes': None
                    })
        
        # Preparar lista de vendedores ordenada por ventas
        lista_vendedores = [
            {
                'id': vid,
                'nombre': vdata['nombre'],
                'total_ventas': vdata['total_ventas'],
                'equipo_id': vdata.get('equipo_id'),
                'equipo_nombre': vdata.get('equipo_nombre', 'Sin equipo')
            }
            for vid, vdata in vendedores_heatmap.items()
        ]
        lista_vendedores.sort(key=lambda x: x['total_ventas'], reverse=True)
        
        # Preparar lista de equipos ordenada por ventas
        lista_equipos = [
            {
                'id': tid,
                'nombre': tdata['nombre'],
                'total_ventas': tdata['total_ventas'],
                'num_miembros': len(tdata['miembros'])
            }
            for tid, tdata in equipos_ventas.items()
        ]
        lista_equipos.sort(key=lambda x: x['total_ventas'], reverse=True)
        
        # Corregir total_ventas_mes si la fuente es Supabase (ya viene filtrado correctamente)
        if data_source == 'supabase':
            # Usar el mismo total que la tarjeta KPI (total_venta ya calculado antes)
            total_ventas_mes = total_venta
            print(f"🔥 Total ventas del mes (heatmap) ajustado desde Supabase: S/ {total_ventas_mes:,.0f}")
        
        print(f"🔥 Heatmap generado: {transacciones_procesadas} transacciones, {celdas_activas} celdas activas")
        print(f"🔥 Total ventas del mes (heatmap): S/ {total_ventas_mes:,.0f} - {len(vendedores_heatmap)} vendedores")
        print(f"🔥 Ventas sin vendedor asignado: {ventas_sin_vendedor}")
        if ventas_excluidas_internacional > 0:
            print(f"⚠️ Ventas excluidas por VENTA INTERNACIONAL: {ventas_excluidas_internacional}")
        if vendedores_heatmap:
            print(f"🔥 Vendedores detectados:")
            for vid, vdata in sorted(vendedores_heatmap.items(), key=lambda x: x[1]['total_ventas'], reverse=True)[:5]:
                print(f"   - {vdata['nombre']}: S/ {vdata['total_ventas']:,.0f}")
        
        # Preparar datos de heatmap por vendedor para el frontend
        heatmap_vendedores_data = {}
        for vendedor_id, matriz in heatmap_por_vendedor.items():
            vendedor_heatmap = []
            matriz_count = heatmap_count_por_vendedor.get(vendedor_id, [[0]*7]*5)
            
            for semana_idx in range(5):
                for dia_idx in range(7):
                    # Usar el mismo mapa de días que creamos antes
                    dia_del_mes = mapa_dias.get((semana_idx, dia_idx))
                    
                    venta_total = matriz[semana_idx][dia_idx]
                    count = matriz_count[semana_idx][dia_idx]
                    
                    if dia_del_mes:
                        vendedor_heatmap.append({
                            'semana': semana_idx,
                            'dia': dia_idx,
                            'valor': venta_total,
                            'transacciones': count,
                            'dia_mes': dia_del_mes
                        })
                    else:
                        # Celda inválida
                        vendedor_heatmap.append({
                            'semana': semana_idx,
                            'dia': dia_idx,
                            'valor': -1,
                            'transacciones': 0,
                            'dia_mes': None
                        })
            
            heatmap_vendedores_data[vendedor_id] = vendedor_heatmap
        
        # --- CLIENTES EN RIESGO ---
        print(f"⚠️ Identificando clientes en riesgo...")
        
        clientes_riesgo = []
        for cliente in clientes_rfm:
            # Clientes en riesgo: sin compras en 60+ días o frecuencia < 1
            if cliente['recency'] >= 60 or cliente['frequency'] < 2:
                nivel_riesgo = 'Alto' if cliente['recency'] >= 90 else 'Medio'
                color_riesgo = '#ff4d4f' if nivel_riesgo == 'Alto' else '#faad14'
                
                clientes_riesgo.append({
                    'cliente': cliente['cliente'],
                    'dias_sin_compra': cliente['recency'],
                    'frecuencia': cliente['frequency'],
                    'valor_historico': cliente['monetary'],
                    'nivel_riesgo': nivel_riesgo,
                    'color': color_riesgo,
                    'categoria_rfm': cliente['categoria'],
                    'canal': cliente.get('canal', 'N/A')  # Agregar canal del cliente
                })
        
        # Ordenar por valor histórico (priorizar clientes valiosos)
        clientes_riesgo_sorted = sorted(clientes_riesgo, key=lambda x: x['valor_historico'], reverse=True)[:20]  # Top 20
        
        print(f"⚠️ Clientes en riesgo identificados: {len(clientes_riesgo_sorted)} de alto valor")
        
        # --- MAPA GEOGRÁFICO DE PENETRACIÓN ---
        print(f"🗺️ Generando análisis geográfico por departamento...")
        
        ventas_por_region = {}
        clientes_por_region = {}
        transacciones_por_region = {}
        
        # Procesar ventas actuales del mes seleccionado
        for sale in sales_data:
            partner_id = sale.get('partner_id')
            balance = sale.get('balance', 0)
            
            # Filtrar ventas internacionales
            partner_name = sale.get('partner_name', '')
            if partner_name == 'VENTA INTERNACIONAL':
                continue
            
            if partner_id and isinstance(partner_id, (list, tuple)) and len(partner_id) > 0:
                partner_id = partner_id[0]
            
            # Obtener state_id del sale (ya viene del partner)
            state_id = sale.get('state_id')
            
            if state_id:
                if isinstance(state_id, (list, tuple)) and len(state_id) > 1:
                    region_name = state_id[1]  # El nombre del departamento está en la posición 1
                else:
                    region_name = "Sin departamento"
            else:
                region_name = "Sin departamento"
            
            # Acumular ventas por región
            if region_name not in ventas_por_region:
                ventas_por_region[region_name] = 0
                clientes_por_region[region_name] = set()
                transacciones_por_region[region_name] = 0
            
            ventas_por_region[region_name] += balance
            if partner_id:
                clientes_por_region[region_name].add(partner_id)
            transacciones_por_region[region_name] += 1
        
        # Convertir sets a counts
        clientes_count_por_region = {k: len(v) for k, v in clientes_por_region.items()}
        
        # Crear lista ordenada de regiones
        datos_geograficos = []
        total_ventas_geo = sum(ventas_por_region.values())
        total_clientes_geo = sum(clientes_count_por_region.values())
        
        for region in ventas_por_region:
            ventas = ventas_por_region[region]
            clientes = clientes_count_por_region[region]
            transacciones = transacciones_por_region[region]
            
            participacion = (ventas / total_ventas_geo * 100) if total_ventas_geo > 0 else 0
            ticket_promedio = (ventas / transacciones) if transacciones > 0 else 0
            
            datos_geograficos.append({
                'region': region,
                'ventas': ventas,
                'clientes': clientes,
                'transacciones': transacciones,
                'participacion': participacion,
                'ticket_promedio': ticket_promedio
            })
        
        # Ordenar por ventas descendente
        datos_geograficos_sorted = sorted(datos_geograficos, key=lambda x: x['ventas'], reverse=True)
        
        print(f"🗺️ Análisis geográfico: {len(datos_geograficos_sorted)} regiones identificadas, S/ {total_ventas_geo:,.2f} en ventas totales")

        # --- ANÁLISIS GEOGRÁFICO DE VENTAS ---
        print("🗺️ Generando análisis geográfico de ventas...")
        ventas_por_departamento = {}
        sales_processed_for_map = 0
        sales_skipped_international = 0
        sales_skipped_non_peru = 0
        sales_skipped_no_state_info = 0

        for sale in sales_data:
            # Excluir ventas internacionales
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                if 'VENTA INTERNACIONAL' in linea_comercial[1].upper():
                    continue
            
            canal_ventas = sale.get('sales_channel_id')
            if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                if 'VENTA INTERNACIONAL' in canal_ventas[1].upper() or 'INTERNACIONAL' in canal_ventas[1].upper():
                    continue

            state_info = sale.get('state_id')
            
            # Lógica mejorada para identificar ventas de Perú
            # Asumimos que si no hay información de país, es una venta local (Perú)
            # Si hay información de país, solo se salta si el nombre del país NO es Perú.
            is_peru_sale = True
            country_info = sale.get('country_id')
            if country_info and isinstance(country_info, list) and len(country_info) > 1:
                country_name = country_info[1].upper()
                if "PERU" not in country_name and "PERÚ" not in country_name:
                    # Es un país explícitamente no-Perú por nombre
                    sales_skipped_non_peru += 1
                    is_peru_sale = False
            
            if not is_peru_sale:
                continue

            if state_info and isinstance(state_info, list) and len(state_info) > 1:
                departamento_nombre_raw = state_info[1]
                
                # --- NORMALIZACIÓN DE NOMBRES DE DEPARTAMENTO ---
                # Convertir a mayúsculas para coincidir con el GeoJSON
                departamento_nombre = departamento_nombre_raw.upper()
                
                # Eliminar sufijos comunes como "(PE)", "(PE )", etc.
                departamento_nombre = re.sub(r'\s*\(PE\)\s*', '', departamento_nombre, flags=re.IGNORECASE).strip()
                
                # Mapeos específicos para corregir discrepancias comunes
                if 'CALLAO' in departamento_nombre:
                    departamento_nombre = 'CALLAO'
                if 'MARTIN' in departamento_nombre:
                    departamento_nombre = 'SAN MARTIN'
                
                # Quitar tildes comunes
                departamento_nombre = departamento_nombre.replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')

                balance = sale.get('balance') or sale.get('price_subtotal', 0)
                if isinstance(balance, str):
                    balance = float(balance.replace(',', ''))
                
                ventas_por_departamento[departamento_nombre] = ventas_por_departamento.get(departamento_nombre, 0) + balance
                sales_processed_for_map += 1
            else:
                sales_skipped_no_state_info += 1

        # Preparar datos para el mapa
        mapa_ventas_data = [{'name': dep, 'value': venta} for dep, venta in ventas_por_departamento.items()]
        print(f"🗺️ Análisis geográfico: {len(mapa_ventas_data)} departamentos con ventas. Total sales processed for map: {sales_processed_for_map}")
        print(f"  Sales skipped (international): {sales_skipped_international}")
        print(f"  Sales skipped (non-Peru): {sales_skipped_non_peru}")
        print(f"  Sales skipped (no state info): {sales_skipped_no_state_info}")

        # --- Procesamiento de datos para gráficos (después del bucle) ---

        # 1. Procesar datos para la tabla principal
        # Generar dinámicamente las líneas comerciales a partir de ventas y metas
        all_lines = {}  # Usar un dict para evitar duplicados, con el id como clave

        # Añadir líneas desde las ventas reales
        for nombre_linea_venta in ventas_por_linea.keys():
            linea_id = nombre_linea_venta.lower().replace(' ', '_')
            all_lines[linea_id] = {'nombre': nombre_linea_venta.upper(), 'id': linea_id}

        # Añadir líneas desde las metas (para aquellas que no tuvieron ventas)
        for linea_id_meta in metas_del_mes.keys():
            # Convertir genvet a terceros si existe en las metas
            if linea_id_meta == 'genvet':
                linea_id_meta = 'terceros'
            
            if linea_id_meta not in all_lines:
                # Reconstruir el nombre desde el ID de la meta
                nombre_reconstruido = linea_id_meta.replace('_', ' ').upper()
                all_lines[linea_id_meta] = {'nombre': nombre_reconstruido, 'id': linea_id_meta}
        
        # Convertir el diccionario de líneas a una lista ordenada por nombre
        lineas_comerciales_dinamicas = sorted(all_lines.values(), key=lambda x: x['nombre'])

        # Excluir líneas no deseadas que pueden venir de los datos
        lineas_a_excluir = ['LICITACION', 'NINGUNO', 'ECOMMERCE', 'GENVET', 'MARCA BLANCA']
        lineas_comerciales_filtradas = [
            linea for linea in lineas_comerciales_dinamicas
            if linea['nombre'].upper() not in lineas_a_excluir
        ]

        # Pre-calcular la venta total para el cálculo de porcentajes
        # IMPORTANTE: Si la fuente es Supabase, usar el total real de get_sales_by_month()
        # porque los datos ya vienen filtrados correctamente
        if data_source == 'supabase':
            # Obtener el total correcto desde Supabase para el mes actual
            meses_es = {
                1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio',
                7: 'julio', 8: 'agosto', 9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
            }
            mes_key = f"{meses_es[mes_sel_int]} {año_sel_int}"
            resumen_mes_actual = supabase_manager.get_sales_by_month(fecha_inicio, fecha_fin)
            total_venta = resumen_mes_actual.get(mes_key, 0)
            print(f"💰 Total venta desde Supabase (ya filtrado): {total_venta:,.2f}")
        else:
            # Para Odoo, usar la suma del loop (que SÍ aplicó filtros)
            total_venta = sum(ventas_por_linea.values())
            print(f"💰 Total venta calculado con filtros Odoo: {total_venta:,.2f}")
        
        total_venta_calculado = total_venta # Renombrar para claridad en el bucle

        # Mes en curso: reflejar la venta real (MTD) recién calculada en la tendencia,
        # porque ese mes aún no está migrado a Supabase (queda en 0 en el resumen).
        # Se hace AQUÍ porque total_venta se finaliza después de construir la tendencia.
        if año_seleccionado == fecha_actual.year and mes_sel_int == fecha_actual.month:
            _mes_curso_key = f"{año_seleccionado}-{fecha_actual.month:02d}"
            for _t in tendencia_12_meses:
                if _t.get('mes') == _mes_curso_key:
                    _t['venta'] = total_venta
                    _t['cumplimiento'] = (total_venta / _t['meta'] * 100) if _t.get('meta') else 0
                    # Marcar como parcial: se muestra en Venta Real pero NO entra en la
                    # regresión de la línea de Tendencia (un mes incompleto la sesgaría).
                    _t['es_parcial'] = True
                    print(f"📈 Tendencia: mes en curso {_mes_curso_key} = S/ {total_venta:,.2f} (venta real MTD, parcial)")
                    break

        # Log de ventas excluidas (solo relevante para Odoo)
        if ventas_categoria_excluida > 0:
            print(f"⚠️ Se excluyeron {ventas_categoria_excluida} líneas por categoría excluida [315, 333, 304, 314, 318, 339]")
        if ventas_sin_linea > 0:
            print(f"⚠️ Se excluyeron {ventas_sin_linea} líneas de venta sin línea comercial")
        if ventas_sin_canal > 0:
            print(f"ℹ️ Se encontraron {ventas_sin_canal} líneas de venta sin canal (pero se procesaron)")
        print(f"💰 Total venta calculado: {total_venta:,.2f}")
        
        print(f"📊 DEBUG: lineas_comerciales_filtradas tiene {len(lineas_comerciales_filtradas)} líneas")
        print(f"📊 DEBUG: ventas_por_linea tiene {len(ventas_por_linea)} líneas")
        if len(ventas_por_linea) > 0:
            print(f"📊 DEBUG: Primeras 3 líneas en ventas_por_linea: {list(ventas_por_linea.keys())[:3]}")

        for linea in lineas_comerciales_filtradas:
            meta = metas_del_mes.get(linea['id'], 0)
            nombre_linea = linea['nombre'].upper()
            
            # Usar ventas reales de Odoo
            venta = ventas_por_linea.get(nombre_linea, 0)
            
            # Usar la meta IPN registrada por el usuario
            meta_pn = metas_ipn_del_mes.get(linea['id'], 0)
            venta_pn = ventas_ipn_por_linea.get(nombre_linea, 0) # Usar el cálculo real de ventas de productos nuevos
            vencimiento = ventas_por_ruta.get(nombre_linea, 0) # Usamos el nuevo cálculo
            
            porcentaje_total = (venta / meta * 100) if meta > 0 else 0
            porcentaje_pn = (venta_pn / meta_pn * 100) if meta_pn > 0 else 0
            porcentaje_sobre_total = (venta / total_venta_calculado * 100) if total_venta_calculado > 0 else 0
            
            # Obtener ventas por canal
            venta_digital = ventas_por_linea_y_canal.get(nombre_linea, {}).get('DIGITAL', 0)
            venta_nacional = ventas_por_linea_y_canal.get(nombre_linea, {}).get('NACIONAL', 0)
            porcentaje_digital = (venta_digital / total_venta_calculado * 100) if total_venta_calculado > 0 else 0
            porcentaje_nacional = (venta_nacional / total_venta_calculado * 100) if total_venta_calculado > 0 else 0

            # Fila TODOS (total)
            datos_lineas.append({
                'nombre': linea['nombre'],
                'meta': meta,
                'venta': venta,
                'porcentaje_total': porcentaje_total,
                'porcentaje_sobre_total': porcentaje_sobre_total,
                'meta_pn': meta_pn,
                'venta_pn': venta_pn,
                'porcentaje_pn': porcentaje_pn,
                'vencimiento_6_meses': vencimiento,
                'canal': 'TODOS'
            })
            
            # Fila DIGITAL
            datos_lineas.append({
                'nombre': linea['nombre'],
                'meta': 0,  # No hay meta por canal
                'venta': venta_digital,
                'porcentaje_total': 0,
                'porcentaje_sobre_total': porcentaje_digital,
                'meta_pn': 0,
                'venta_pn': 0,
                'porcentaje_pn': 0,
                'vencimiento_6_meses': 0,
                'canal': 'DIGITAL'
            })
            
            # Fila NACIONAL
            datos_lineas.append({
                'nombre': linea['nombre'],
                'meta': 0,  # No hay meta por canal
                'venta': venta_nacional,
                'porcentaje_total': 0,
                'porcentaje_sobre_total': porcentaje_nacional,
                'meta_pn': 0,
                'venta_pn': 0,
                'porcentaje_pn': 0,
                'vencimiento_6_meses': 0,
                'canal': 'NACIONAL'
            })
            
            # Los totales de metas ya se calcularon. Aquí solo sumamos los totales de ventas.
            total_venta_pn += venta_pn
            total_vencimiento += vencimiento
        # --- 2. Calcular KPIs ---
        # Días laborables restantes (Lunes a Sábado)
        dias_restantes = 0
        ritmo_diario_requerido = 0
        if mes_seleccionado == fecha_actual.strftime('%Y-%m'):
            hoy = fecha_actual.day
            ultimo_dia_mes = calendar.monthrange(año_actual, fecha_actual.month)[1]
            for dia in range(hoy, ultimo_dia_mes + 1):
                # weekday() -> Lunes=0, Domingo=6
                if datetime(año_actual, fecha_actual.month, dia).weekday() < 6:
                    dias_restantes += 1
            
            porcentaje_restante = 100 - ((total_venta / total_meta * 100) if total_meta > 0 else 100)
            if porcentaje_restante > 0 and dias_restantes > 0:
                ritmo_diario_requerido = porcentaje_restante / dias_restantes

        # Calcular KPIs
        kpis = {
            'meta_total': total_meta,
            'venta_total': total_venta,
            'porcentaje_avance': (total_venta / total_meta * 100) if total_meta > 0 else 0,
            'meta_ipn': total_meta_pn,
            'venta_ipn': total_venta_pn,
            'porcentaje_avance_ipn': (total_venta_pn / total_meta_pn * 100) if total_meta_pn > 0 else 0,
            'vencimiento_6_meses': total_vencimiento,
            'avance_diario_total': ((total_venta / total_meta * 100) / dia_actual) if total_meta > 0 and dia_actual > 0 else 0,
            'avance_diario_ipn': ((total_venta_pn / total_meta_pn * 100) / dia_actual) if total_meta_pn > 0 and dia_actual > 0 else 0,
            'ritmo_diario_requerido': ritmo_diario_requerido,
            'total_clientes_cartera': total_clientes,
            'clientes_activos': num_clientes_activos,
            'cobertura_clientes': cobertura_clientes
        }

        # --- KPIs ANUALES (acumulado del año) ---
        # Venta y meta anuales = suma de la tendencia (meses cerrados de Supabase
        # + mes en curso en vivo ya inyectado). Total = el "Total {año}" del gráfico.
        venta_total_anual = sum(t.get('venta', 0) for t in tendencia_12_meses)
        meta_total_anual = sum(t.get('meta', 0) for t in tendencia_12_meses)

        # Meta IPN anual = suma de metas_ipn de los 12 meses
        meta_ipn_anual = 0
        for _m in range(1, 13):
            _mk = f"{año_seleccionado}-{_m:02d}"
            meta_ipn_anual += sum(metas_historicas.get(_mk, {}).get('metas_ipn', {}).values())

        # Venta IPN anual = IPN de meses cerrados (Supabase) + mes en curso (vivo)
        venta_ipn_anual = 0
        try:
            if SUPABASE_ENABLED:
                venta_ipn_anual = supabase_manager.get_ipn_total(
                    f"{año_seleccionado}-01-01", f"{año_seleccionado}-12-31")
        except Exception as _e:
            print(f"⚠️ Error calculando Venta IPN anual: {_e}")
        # Sumar el mes en curso (no migrado) solo si lo estamos viendo, para que
        # cuadre con venta_total_anual (que también incluye el mes en curso solo entonces)
        if año_seleccionado == fecha_actual.year and mes_sel_int == fecha_actual.month:
            venta_ipn_anual += total_venta_pn

        kpis_anual = {
            'meta_total': meta_total_anual,
            'venta_total': venta_total_anual,
            'porcentaje_avance': (venta_total_anual / meta_total_anual * 100) if meta_total_anual > 0 else 0,
            'meta_ipn': meta_ipn_anual,
            'venta_ipn': venta_ipn_anual,
            'porcentaje_avance_ipn': (venta_ipn_anual / meta_ipn_anual * 100) if meta_ipn_anual > 0 else 0,
        }

        # --- Avance lineal: proyección de cierre y faltante ---
        # Proyección mensual lineal: proyectar ventas actuales al mes completo
        try:
            dias_en_mes = calendar.monthrange(int(año_sel), int(mes_sel))[1]
        except Exception:
            dias_en_mes = 30

        if dia_actual > 0:
            proyeccion_mensual = (total_venta / dia_actual) * dias_en_mes
        else:
            proyeccion_mensual = 0

        avance_lineal_pct = (proyeccion_mensual / total_meta * 100) if total_meta > 0 else 0
        faltante_meta = max(total_meta - total_venta, 0)

        # Cálculos específicos para IPN (usando las variables ya calculadas)
        # total_meta_pn ya está calculado arriba
        # total_venta_pn ya está calculado arriba
        
        # Proyección lineal IPN
        if dia_actual > 0:
            promedio_diario_ipn = total_venta_pn / dia_actual
            proyeccion_mensual_ipn = promedio_diario_ipn * dias_en_mes
        else:
            proyeccion_mensual_ipn = 0

        avance_lineal_ipn_pct = (proyeccion_mensual_ipn / total_meta_pn * 100) if total_meta_pn > 0 else 0
        faltante_meta_ipn = max(total_meta_pn - total_venta_pn, 0)

        
        # 3. Ordenar productos para el gráfico Top 7
        # Ordenar productos por ventas y tomar los top 7
        productos_ordenados = sorted(ventas_por_producto.items(), key=lambda x: x[1], reverse=True)[:7]
        
        datos_productos = []
        for nombre_producto, venta in productos_ordenados:
            datos_productos.append({
                'nombre': nombre_producto,
                'venta': venta,
                'ciclo_vida': ciclo_vida_por_producto.get(nombre_producto, 'No definido')
            })
        
        # 4. Ordenar datos para el gráfico de Ciclo de Vida
        # Convertir a lista ordenada por ventas
        datos_ciclo_vida = []
        for ciclo, venta in sorted(ventas_por_ciclo_vida.items(), key=lambda x: x[1], reverse=True):
            datos_ciclo_vida.append({
                'ciclo': ciclo,
                'venta': venta
            })
        
        # --- INICIO: LÓGICA PARA LA TABLA DEL EQUIPO ECOMMERCE ---
        datos_ecommerce = []
        kpis_ecommerce = {'meta_total': 0, 'venta_total': 0, 'porcentaje_avance': 0}

        # 1. Obtener miembros y metas del equipo ECOMMERCE
        equipos_guardados = gs_manager.read_equipos()        
        ecommerce_vendor_ids = {str(vid) for vid in equipos_guardados.get('ecommerce', [])}
        
        if ecommerce_vendor_ids:
            # 2. Obtener la meta total de ECOMMERCE desde las metas por línea
            meta_ecommerce = metas_del_mes.get('ecommerce', 0)
            kpis_ecommerce['meta_total'] = meta_ecommerce

            # 3. Calcular ventas del equipo ECOMMERCE, agrupadas por LÍNEA COMERCIAL
            ventas_por_linea_ecommerce = {}
            for sale in sales_data:
                user_info = sale.get('invoice_user_id')
                if user_info and isinstance(user_info, list) and len(user_info) > 1:
                    vendedor_id = str(user_info[0])
                    # Si la venta pertenece a un vendedor de ECOMMERCE
                    if vendedor_id in ecommerce_vendor_ids:
                        balance = float(sale.get('balance', 0))
                        
                        # Agrupar por línea comercial con normalización
                        linea_info = sale.get('commercial_line_national_id')
                        linea_nombre = 'N/A'
                        if linea_info and isinstance(linea_info, list) and len(linea_info) > 1:
                            linea_nombre_original = linea_info[1].upper()
                            # Aplicar normalización para agrupar GENVET y MARCA BLANCA como TERCEROS
                            linea_nombre = normalizar_linea_comercial(linea_nombre_original)
                        
                        ventas_por_linea_ecommerce[linea_nombre] = ventas_por_linea_ecommerce.get(linea_nombre, 0) + balance

            # 4. Construir la tabla de datos para la plantilla
            for linea, venta in ventas_por_linea_ecommerce.items():
                datos_ecommerce.append({
                    'nombre': linea, # Ahora es el nombre de la línea comercial
                    'venta': venta
                })
                kpis_ecommerce['venta_total'] += venta

            # 5. Calcular el porcentaje de avance total del equipo
            if kpis_ecommerce['meta_total'] > 0:
                kpis_ecommerce['porcentaje_avance'] = (kpis_ecommerce['venta_total'] / kpis_ecommerce['meta_total']) * 100

            # 6. Calcular el porcentaje de participación de cada línea sobre el total del equipo
            if kpis_ecommerce['venta_total'] > 0:
                for linea_data in datos_ecommerce:
                    linea_data['porcentaje_sobre_total'] = (linea_data['venta'] / kpis_ecommerce['venta_total']) * 100
            else:
                for linea_data in datos_ecommerce:
                    linea_data['porcentaje_sobre_total'] = 0

            # Ordenar las líneas por venta descendente
            datos_ecommerce = sorted(datos_ecommerce, key=lambda x: x['venta'], reverse=True)

        # --- FIN: LÓGICA PARA LA TABLA DEL EQUIPO ECOMMERCE ---

        # --- INICIO: GRÁFICO DE VENTAS POR MES CON FILTROS ---
        # Render Free: red ultra-lenta (0.1 CPU) hace IMPOSIBLE transferir 31K registros
        # - Odoo: >5 min para 4K registros (XML-RPC lento)
        # - Supabase: >5 min para 31K registros (32 requests HTTP de 1K c/u)
        # SOLUCIÓN: Deshabilitar en producción, habilitar solo en desarrollo local
        
        if is_render:
            # Render Free: Omitir siempre (red demasiado lenta)
            print(f"⏭️  Gráfico de productos DESHABILITADO en Render Free Tier")
            print(f"    Motivo: Transferir {31982 if data_source == 'supabase' else '4K+'} registros sobre red 0.1 CPU = timeout")
            datos_ventas_mes_filtros = {
                'datos': [],
                'filtros': {
                    'lineas_comerciales': [],
                    'categorias': [],
                    'ciclos_vida': [],
                    'vias_administracion': [],
                    'clasificaciones': [],
                    'formas_farmaceuticas': [],
                    'lineas_produccion': []
                }
            }
        else:
            # Desarrollo local: Red rápida, generar normalmente
            print(f"📊 Generando gráfico de productos filtrados para año {año_seleccionado} ({data_source})")
            datos_ventas_mes_filtros = generar_datos_ventas_mes(
                año_seleccionado,
                data_source,
                fecha_actual,
                sales_data_anual_override=sales_data_anual_compartida
            )
        # --- FIN: GRÁFICO DE VENTAS POR MES CON FILTROS ---

        # Ordenar los datos de la tabla: primero las filas TODOS, luego DIGITAL, luego NACIONAL
        # Ordenar por venta descendente dentro de cada grupo
        datos_todos = [d for d in datos_lineas if d.get('canal') == 'TODOS']
        datos_digital = [d for d in datos_lineas if d.get('canal') == 'DIGITAL']
        datos_nacional = [d for d in datos_lineas if d.get('canal') == 'NACIONAL']
        
        datos_todos_sorted = sorted(datos_todos, key=lambda x: x['venta'], reverse=True)
        datos_digital_sorted = sorted(datos_digital, key=lambda x: x['venta'], reverse=True)
        datos_nacional_sorted = sorted(datos_nacional, key=lambda x: x['venta'], reverse=True)
        datos_digital_por_linea = {item['nombre']: item for item in datos_digital_sorted}
        datos_nacional_por_linea = {item['nombre']: item for item in datos_nacional_sorted}
        
        # Intercalar: para cada línea TODOS, agregar su DIGITAL y NACIONAL correspondiente
        datos_lineas_tabla_sorted = []
        for linea_todos in datos_todos_sorted:
            nombre_linea = linea_todos['nombre']
            # Agregar la fila TODOS
            datos_lineas_tabla_sorted.append(linea_todos)
            # Agregar la fila DIGITAL correspondiente
            linea_digital = datos_digital_por_linea.get(nombre_linea)
            if linea_digital:
                datos_lineas_tabla_sorted.append(linea_digital)
            # Agregar la fila NACIONAL correspondiente
            linea_nacional = datos_nacional_por_linea.get(nombre_linea)
            if linea_nacional:
                datos_lineas_tabla_sorted.append(linea_nacional)

        # Preparar los datos para renderizar
        render_data = {
            'meses_disponibles': meses_disponibles,
            'años_disponibles': años_disponibles,
            'año_seleccionado': año_seleccionado,
            'año_actual': año_actual,
            'mes_seleccionado': mes_seleccionado,
            'mes_nombre': mes_nombre,
            'dia_actual': dia_actual,
            'kpis': kpis,
            'kpis_anual': kpis_anual,
            'datos_lineas': datos_lineas,
            'datos_lineas_tabla': datos_lineas_tabla_sorted,
            'datos_clientes_por_linea': datos_clientes_por_linea,
            'datos_cobertura_canal': datos_cobertura_canal,
            'datos_cobertura_grupos': datos_cobertura_grupos,
            'cobertura_clientes': cobertura_clientes,  # Agregar cobertura general
            'total_clientes': total_clientes,  # Agregar cartera total
            'num_clientes_activos': num_clientes_activos,  # Agregar clientes activos
            'datos_frecuencia_linea': datos_frecuencia_linea,
            'clientes_rfm': clientes_rfm_sorted[:100],  # Top 100 clientes
            'segmentos_rfm': segmentos_rfm,
            'segmentos_por_canal': segmentos_por_canal,  # Nuevo: Segmentos RFM por canal
            'tendencia_12_meses': tendencia_12_meses,
            'clientes_riesgo': clientes_riesgo_sorted,
            'heatmap_ventas': heatmap_ventas,
            'heatmap_dias': dias_labels,
            'heatmap_semanas': semanas_labels,
            'heatmap_vendedores': lista_vendedores,  # Lista de vendedores del mes
            'heatmap_por_vendedor': heatmap_vendedores_data,  # Datos por vendedor
            'heatmap_equipos': lista_equipos,  # Lista de equipos de ventas
            'mapa_ventas_data': mapa_ventas_data,
            'datos_geograficos': datos_geograficos_sorted,  # Nuevo: Mapa geográfico
            'datos_productos': datos_productos,
            'datos_ciclo_vida': datos_ciclo_vida if 'datos_ciclo_vida' in locals() else [],
            'datos_ventas_mes_filtros': datos_ventas_mes_filtros,  # Nuevo: Gráfico de ventas por mes
            'fecha_actual': fecha_actual,
            'avance_lineal_pct': avance_lineal_pct,
            'faltante_meta': faltante_meta,
            'avance_lineal_ipn_pct': avance_lineal_ipn_pct,
            'faltante_meta_ipn': faltante_meta_ipn,
            'datos_ecommerce': datos_ecommerce,
            'kpis_ecommerce': kpis_ecommerce,
            'is_admin': is_admin,
            'desde_cache': False,  # Datos frescos
            'grupos_venta': grupos_venta  # Nuevo: Grupos de venta para filtro del mapa
        }
        
        # Guardar en caché para futuras solicitudes.
        # La nueva lógica de caché maneja la expiración para el mes actual.
        cache_data = render_data.copy()
        cache_data.pop('is_admin', None)  # No cachear datos de sesión
        cache_data['desde_cache'] = False  # Se establecerá en True al leer del caché
        save_to_cache(año_sel_int, mes_sel_int, cache_data)

        # DEBUG: Verificar qué se pasa al template
        print(f"\n{'='*80}")
        print(f"🎯 RENDER TEMPLATE - DATOS FRESCOS")
        print(f"   año_seleccionado: {render_data['año_seleccionado']}")
        print(f"   mes_seleccionado: {render_data['mes_seleccionado']}")
        print(f"   mes_nombre: {render_data['mes_nombre']}")
        print(f"   tendencia_12_meses: {len(render_data['tendencia_12_meses'])} meses")
        if len(render_data['tendencia_12_meses']) > 0:
            print(f"   Primer mes tendencia: {render_data['tendencia_12_meses'][0].get('mes', 'N/A')}")
        print(f"{'='*80}\n")
        
        return render_template('dashboard_clean.html', **render_data)
    
    except Exception as e:
        print(f"\n❌ ERROR EN DASHBOARD: {str(e)}")
        print(f"   Tipo de error: {type(e).__name__}")
        import traceback
        print(f"   Traceback:")
        traceback.print_exc()
        print(f"\n")
        flash(f'Error al obtener datos del dashboard: {str(e)}', 'danger')
        
        # Crear datos por defecto para evitar errores
        fecha_actual = datetime.now()
        año_actual = fecha_actual.year
        kpis_default = {
            'meta_total': 0,
            'venta_total': 0,
            'porcentaje_avance': 0,
            'meta_ipn': 0,
            'venta_ipn': 0,
            'porcentaje_avance_ipn': 0,
            'vencimiento_6_meses': 0,
            'avance_diario_total': 0,
            'avance_diario_ipn': 0,
            'total_clientes_cartera': 0,
            'clientes_activos': 0,
            'cobertura_clientes': 0
        }
        
        return render_template('dashboard_clean.html',
                             meses_disponibles=[{
                                 'key': fecha_actual.strftime('%Y-%m'),
                                 'nombre': f"{fecha_actual.strftime('%B')} {fecha_actual.year}"
                             }],
                             años_disponibles=list(range(2020, año_actual + 1)),
                             año_seleccionado=año_actual,
                             año_actual=año_actual,
                             mes_seleccionado=fecha_actual.strftime('%Y-%m'),
                             mes_nombre=f"{fecha_actual.strftime('%B').upper()} {fecha_actual.year}",
                             dia_actual=fecha_actual.day,
                             kpis=kpis_default,
                             datos_lineas=[], # Se mantiene vacío en caso de error
                             datos_lineas_tabla=[],
                             datos_clientes_por_linea=[], # Nueva tabla vacía en caso de error
                             datos_cobertura_canal=[],
                             datos_frecuencia_linea=[],
                             clientes_rfm=[],
                             segmentos_rfm={},
                             tendencia_12_meses=[],
                             clientes_riesgo=[],
                             heatmap_ventas=[],
                             heatmap_dias=['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom'],
                             heatmap_semanas=['Semana 1', 'Semana 2', 'Semana 3', 'Semana 4', 'Semana 5'],
                             mapa_ventas_data=[],
                             datos_productos=[],
                             datos_ciclo_vida=[],
                             datos_ventas_mes_filtros={'registros': [], 'filtros': {}},  # Vacío en caso de error
                             fecha_actual=fecha_actual,
                             avance_lineal_pct=0,
                             faltante_meta=0,
                             datos_ecommerce=[],
                             kpis_ecommerce={},
                             is_admin=is_admin) # Pasar el flag a la plantilla


@app.route('/dashboard_linea')
def dashboard_linea():
    if 'username' not in session:
        return redirect(url_for('login'))

    # --- Lógica de Permisos de Administrador ---
    is_admin = session.get('username') in ADMIN_USERS

    try:
        # --- 1. OBTENER FILTROS ---
        fecha_actual = datetime.now()
        mes_seleccionado = request.args.get('mes', fecha_actual.strftime('%Y-%m'))
        año_actual = fecha_actual.year
        meses_disponibles = get_meses_del_año(año_actual)

        linea_seleccionada_nombre = request.args.get('linea_nombre', 'PETMEDICA') # Default a PETMEDICA si no se especifica

        # --- NUEVA LÓGICA DE FILTRADO POR DÍA ---
        dia_fin_param = request.args.get('dia_fin')
        año_sel, mes_sel = mes_seleccionado.split('-')

        if dia_fin_param:
            try:
                dia_actual = int(dia_fin_param)
                fecha_fin = f"{año_sel}-{mes_sel}-{str(dia_actual).zfill(2)}"
            except (ValueError, TypeError):
                dia_fin_param = None
        
        if not dia_fin_param:
            if mes_seleccionado == fecha_actual.strftime('%Y-%m'):
                dia_actual = fecha_actual.day
            else:
                ultimo_dia_mes = calendar.monthrange(int(año_sel), int(mes_sel))[1]
                dia_actual = ultimo_dia_mes
            fecha_fin = f"{año_sel}-{mes_sel}-{str(dia_actual).zfill(2)}"
        
        fecha_inicio = f"{año_sel}-{mes_sel}-01"
        # --- FIN DE LA NUEVA LÓGICA ---

        # Mapeo de nombre de línea a ID para cargar metas
        mapeo_nombre_a_id = {
            'PETMEDICA': 'petmedica', 'AGROVET': 'agrovet', 'PET NUTRISCIENCE': 'pet_nutriscience',
            'AVIVET': 'avivet', 'OTROS': 'otros',
            'TERCEROS': 'terceros', 'INTERPET': 'interpet',
        }
        linea_seleccionada_id = mapeo_nombre_a_id.get(linea_seleccionada_nombre.upper(), 'petmedica')

        # --- 2. OBTENER DATOS ---
        # fecha_inicio y fecha_fin se calculan arriba usando la lógica de dia_fin.
        # Asegurar que fecha_inicio siempre esté definida
        año_sel, mes_sel = mes_seleccionado.split('-')
        fecha_inicio = f"{año_sel}-{mes_sel}-01"
        # Si no se definió fecha_fin arriba (por alguna razón), usar el último día del mes
        if 'fecha_fin' not in locals():
            ultimo_dia = calendar.monthrange(int(año_sel), int(mes_sel))[1]
            fecha_fin = f"{año_sel}-{mes_sel}-{ultimo_dia}"

        # Cargar metas de vendedores para el mes y línea seleccionados
        # La estructura es metas[equipo_id][vendedor_id][mes_key]
        metas_vendedores_historicas = gs_manager.read_metas()
        # 1. Obtener todas las metas del equipo/línea
        metas_del_equipo = metas_vendedores_historicas.get(linea_seleccionada_id, {})

        # Obtener todos los vendedores de Odoo
        todos_los_vendedores = {str(v['id']): v['name'] for v in data_manager.get_all_sellers()}

        # Obtener ventas del mes
        sales_data = data_manager.get_sales_lines(
            date_from=fecha_inicio,
            date_to=fecha_fin,
            limit=10000
        )

        # --- PRE-FILTRAR VENTAS INTERNACIONALES PARA EFICIENCIA ---
        sales_data_processed = []
        for sale in sales_data:
            # Excluir VENTA INTERNACIONAL (exportaciones) por línea comercial
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                if 'VENTA INTERNACIONAL' in linea_comercial[1].upper():
                    continue
            
            # Excluir VENTA INTERNACIONAL por canal de ventas
            canal_ventas = sale.get('sales_channel_id')
            if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                nombre_canal = canal_ventas[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_canal or 'INTERNACIONAL' in nombre_canal:
                    continue
            
            sales_data_processed.append(sale)

        # --- 3. PROCESAR Y AGREGAR DATOS POR VENDEDOR ---
        ventas_por_vendedor = {}
        ventas_ipn_por_vendedor = {}
        ventas_vencimiento_por_vendedor = {}
        ventas_por_producto = {}
        ventas_por_ciclo_vida = {}
        ventas_por_forma = {}
        ajustes_sin_vendedor = 0 # Para notas de crédito sin vendedor
        nombres_vendedores_con_ventas = {} # BUGFIX: Guardar nombres de vendedores con ventas

        for sale in sales_data_processed: # Usar los datos pre-filtrados
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                nombre_linea_original = linea_comercial[1].upper()
                # Aplicar normalización para agrupar GENVET y MARCA BLANCA como TERCEROS
                nombre_linea_actual = normalizar_linea_comercial(nombre_linea_original)

                # Filtrar por la línea comercial seleccionada
                if nombre_linea_actual == linea_seleccionada_nombre.upper():
                    balance = float(sale.get('balance', 0))
                    user_info = sale.get('invoice_user_id')

                    # Si hay un vendedor asignado, se procesa normalmente
                    if user_info and isinstance(user_info, list) and len(user_info) > 1:
                        vendedor_id = str(user_info[0])
                        nombres_vendedores_con_ventas[vendedor_id] = user_info[1] # Guardar el nombre

                        # Agrupar ventas totales
                        ventas_por_vendedor[vendedor_id] = ventas_por_vendedor.get(vendedor_id, 0) + balance

                        # Agrupar ventas IPN
                        if sale.get('product_life_cycle') == 'nuevo':
                            ventas_ipn_por_vendedor[vendedor_id] = ventas_ipn_por_vendedor.get(vendedor_id, 0) + balance
                        
                        # Agrupar ventas por vencimiento < 6 meses
                        ruta = sale.get('route_id')
                        if isinstance(ruta, list) and len(ruta) > 0 and ruta[0] in [18, 19]:
                            ventas_vencimiento_por_vendedor[vendedor_id] = ventas_vencimiento_por_vendedor.get(vendedor_id, 0) + balance
                    
                    # Si NO hay vendedor, se agrupa como un ajuste (ej. Nota de Crédito)
                    else:
                        ajustes_sin_vendedor += balance

                    # Agrupar para gráficos (Top Productos, Ciclo Vida, Forma Farmacéutica)
                    # Esto se hace para todas las transacciones de la línea, con o sin vendedor
                    producto_nombre = sale.get('name', '').strip()
                    if producto_nombre:
                        # Limpiar nombres de ATREVIA eliminando indicadores de tamaño/presentación
                        producto_nombre_limpio = limpiar_nombre_atrevia(producto_nombre)
                        ventas_por_producto[producto_nombre_limpio] = ventas_por_producto.get(producto_nombre_limpio, 0) + balance

                    ciclo_vida = sale.get('product_life_cycle', 'No definido')
                    ventas_por_ciclo_vida[ciclo_vida] = ventas_por_ciclo_vida.get(ciclo_vida, 0) + balance

                    forma_farma = sale.get('pharmaceutical_forms_id')
                    nombre_forma = forma_farma[1] if forma_farma and len(forma_farma) > 1 else 'Instrumental'
                    ventas_por_forma[nombre_forma] = ventas_por_forma.get(nombre_forma, 0) + balance

        # --- 4. CONSTRUIR ESTRUCTURA DE DATOS PARA LA PLANTILLA ---
        datos_vendedores = []
        total_meta = 0
        total_venta = 0
        total_meta_ipn = 0
        total_venta_ipn = 0
        total_vencimiento = 0

        # --- 4.1. UNIFICAR VENDEDORES ---
        # Combinar los vendedores oficiales del equipo con los que tuvieron ventas reales en la línea.
        # Esto asegura que mostremos a todos los miembros del equipo (incluso con 0 ventas)
        # y también a cualquier otra persona que haya vendido en esta línea sin ser miembro oficial.
        equipos_guardados = gs_manager.read_equipos()
        miembros_oficiales_ids = {str(vid) for vid in equipos_guardados.get(linea_seleccionada_id, [])}
        vendedores_con_ventas_ids = set(ventas_por_vendedor.keys())
        
        todos_los_vendedores_a_mostrar_ids = sorted(list(miembros_oficiales_ids | vendedores_con_ventas_ids))

        # --- 4.2. CONSTRUIR LA TABLA DE VENDEDORES ---
        for vendedor_id in todos_los_vendedores_a_mostrar_ids:
            # BUGFIX: Priorizar el nombre de la venta, luego la lista general, y como último recurso el ID.
            vendedor_nombre = nombres_vendedores_con_ventas.get(vendedor_id, 
                                todos_los_vendedores.get(vendedor_id, f"Vendedor ID {vendedor_id}"))

            
            # Obtener ventas (será 0 si es un miembro oficial sin ventas)
            venta = ventas_por_vendedor.get(vendedor_id, 0)
            venta_ipn = ventas_ipn_por_vendedor.get(vendedor_id, 0)
            vencimiento = ventas_vencimiento_por_vendedor.get(vendedor_id, 0)

            # Asignar meta SOLO si el vendedor es un miembro oficial del equipo
            meta = 0
            meta_ipn = 0
            if vendedor_id in miembros_oficiales_ids:
                meta_guardada = metas_del_equipo.get(vendedor_id, {}).get(mes_seleccionado, {})
                meta = float(meta_guardada.get('meta', 0))
                meta_ipn = float(meta_guardada.get('meta_ipn', 0))

            # Añadir la fila del vendedor a la tabla
            datos_vendedores.append({
                'id': vendedor_id,
                'nombre': vendedor_nombre,
                'meta': meta,
                'venta': venta,
                'porcentaje_avance': (venta / meta * 100) if meta > 0 else 0,
                'meta_ipn': meta_ipn,
                'venta_ipn': venta_ipn,
                'porcentaje_avance_ipn': (venta_ipn / meta_ipn * 100) if meta_ipn > 0 else 0,
                'vencimiento_6_meses': vencimiento
            })

            # Sumar a los totales generales de la línea.
            # La meta solo se suma si fue asignada (es decir, si es miembro oficial).
            # La venta se suma siempre.
            total_meta += meta
            total_venta += venta
            total_meta_ipn += meta_ipn
            total_venta_ipn += venta_ipn
            total_vencimiento += vencimiento

        # --- 4.3. AÑADIR AJUSTES SIN VENDEDOR ---
        if ajustes_sin_vendedor != 0:
            datos_vendedores.append({
                'id': 'ajustes',
                'nombre': 'Ajustes y Notas de Crédito (Sin Vendedor)',
                'meta': 0, 'venta': ajustes_sin_vendedor, 'porcentaje_avance': 0,
                'meta_ipn': 0, 'venta_ipn': 0, 'porcentaje_avance_ipn': 0,
                'vencimiento_6_meses': 0
            })
            # Sumar los ajustes al total de ventas de la línea
            total_venta += ajustes_sin_vendedor

        # Añadir porcentaje sobre el total a cada vendedor
        if total_venta > 0:
            for v in datos_vendedores:
                v['porcentaje_sobre_total'] = (v.get('venta', 0) / total_venta) * 100
        else:
            for v in datos_vendedores:
                v['porcentaje_sobre_total'] = 0

        # --- 4.4. FILTRAR VENDEDORES CON VENTA NEGATIVA ---
        # Si un vendedor solo tiene notas de crédito (venta < 0), no se muestra en la tabla,
        # pero su valor ya fue sumado (restado) al total_venta para mantener la consistencia.
        datos_vendedores_final = [v for v in datos_vendedores if v['venta'] >= 0 or v['id'] == 'ajustes']

        # Ordenar por venta descendente
        datos_vendedores_final = sorted(datos_vendedores_final, key=lambda x: x['venta'], reverse=True)

        # --- 5. CALCULAR KPIs DE LÍNEA ---
        ritmo_diario_requerido_linea = 0
        if mes_seleccionado == fecha_actual.strftime('%Y-%m'):
            hoy = fecha_actual.day
            ultimo_dia_mes = calendar.monthrange(año_actual, fecha_actual.month)[1]
            dias_restantes = 0
            for dia in range(hoy, ultimo_dia_mes + 1):
                if datetime(año_actual, fecha_actual.month, dia).weekday() < 6: # L-S
                    dias_restantes += 1
            
            porcentaje_restante = 100 - ((total_venta / total_meta * 100) if total_meta > 0 else 100)
            if porcentaje_restante > 0 and dias_restantes > 0:
                ritmo_diario_requerido_linea = porcentaje_restante / dias_restantes

        # KPIs generales para la línea
        kpis = {
            'meta_total': total_meta,
            'venta_total': total_venta,
            'porcentaje_avance': (total_venta / total_meta * 100) if total_meta > 0 else 0,
            'meta_ipn': total_meta_ipn,
            'venta_ipn': total_venta_ipn,
            'porcentaje_avance_ipn': (total_venta_ipn / total_meta_ipn * 100) if total_meta_ipn > 0 else 0,
            'vencimiento_6_meses': total_vencimiento,
            'avance_diario_total': ((total_venta / total_meta * 100) / dia_actual) if total_meta > 0 and dia_actual > 0 else 0,
            'avance_diario_ipn': ((total_venta_ipn / total_meta_ipn * 100) / dia_actual) if total_meta_ipn > 0 and dia_actual > 0 else 0,
            'ritmo_diario_requerido': ritmo_diario_requerido_linea
        }

        # --- Avance lineal específico de la línea: proyección de cierre y faltante ---
        try:
            dias_en_mes = calendar.monthrange(int(año_sel), int(mes_sel))[1]
        except Exception:
            dias_en_mes = 30

        if dia_actual > 0:
            proyeccion_mensual_linea = (total_venta / dia_actual) * dias_en_mes
        else:
            proyeccion_mensual_linea = 0

        avance_lineal_pct = (proyeccion_mensual_linea / total_meta * 100) if total_meta > 0 else 0
        faltante_meta = max(total_meta - total_venta, 0)

        # Cálculos específicos para IPN de la línea
        if dia_actual > 0:
            promedio_diario_ipn_linea = total_venta_ipn / dia_actual
            proyeccion_mensual_ipn_linea = promedio_diario_ipn_linea * dias_en_mes
        else:
            proyeccion_mensual_ipn_linea = 0

        avance_lineal_ipn_pct = (proyeccion_mensual_ipn_linea / total_meta_ipn * 100) if total_meta_ipn > 0 else 0
        faltante_meta_ipn = max(total_meta_ipn - total_venta_ipn, 0)

        # Datos para gráficos
        productos_ordenados = sorted(ventas_por_producto.items(), key=lambda x: x[1], reverse=True)[:7]
        datos_productos = [{'nombre': n, 'venta': v} for n, v in productos_ordenados]

        datos_ciclo_vida = [{'ciclo': c, 'venta': v} for c, v in ventas_por_ciclo_vida.items()]
        datos_forma_farmaceutica = [{'forma': f, 'venta': v} for f, v in ventas_por_forma.items()]

        # --- LÓGICA MEJORADA PARA OBTENER LÍNEAS COMERCIALES DISPONIBLES ---
        # Replicar la misma lógica del dashboard principal para consistencia.
        
        # 1. Obtener metas del mes para incluir líneas con metas pero sin ventas.
        año_sel_int = int(año_sel)
        metas_historicas = load_metas(año_sel_int)
        metas_del_mes = metas_historicas.get(mes_seleccionado, {}).get('metas', {})
        
        # 2. Unificar líneas desde ventas y metas.
        all_lines_dict = {}

        # Desde ventas (aplicando normalización)
        for sale in sales_data_processed: # Usar datos ya filtrados de ventas internacionales
            linea_obj = sale.get('commercial_line_national_id')
            if linea_obj and isinstance(linea_obj, list) and len(linea_obj) > 1:
                linea_nombre_original = linea_obj[1].upper()
                # Aplicar normalización para agrupar GENVET y MARCA BLANCA como TERCEROS
                linea_nombre = normalizar_linea_comercial(linea_nombre_original)
                if linea_nombre not in all_lines_dict:
                    all_lines_dict[linea_nombre] = linea_nombre

        # Desde metas
        for linea_id_meta in metas_del_mes.keys():
            nombre_linea_meta = linea_id_meta.replace('_', ' ').upper()
            if nombre_linea_meta not in all_lines_dict:
                all_lines_dict[nombre_linea_meta] = nombre_linea_meta

        # 3. Filtrar y ordenar
        lineas_a_excluir = ['LICITACION', 'NINGUNO', 'ECOMMERCE', 'VENTA INTERNACIONAL']
        lineas_disponibles = sorted([nombre for nombre in all_lines_dict.values() if nombre not in lineas_a_excluir])
        # --- FIN DE LA LÓGICA MEJORADA ---
        return render_template('dashboard_linea.html',
                               linea_nombre=linea_seleccionada_nombre,
                               mes_seleccionado=mes_seleccionado,
                               meses_disponibles=meses_disponibles,
                               kpis=kpis,
                               datos_vendedores=datos_vendedores_final,
                               datos_productos=datos_productos,
                               datos_ciclo_vida=datos_ciclo_vida,
                               datos_forma_farmaceutica=datos_forma_farmaceutica,
                               lineas_disponibles=lineas_disponibles,
                               fecha_actual=fecha_actual,
                               dia_actual=dia_actual,
                               avance_lineal_pct=avance_lineal_pct,
                               faltante_meta=faltante_meta,
                               avance_lineal_ipn_pct=avance_lineal_ipn_pct,
                               faltante_meta_ipn=faltante_meta_ipn,
                               is_admin=is_admin) # Pasar el flag a la plantilla

    except Exception as e:
        flash(f'Error al generar el dashboard para la línea: {str(e)}', 'danger')
        # En caso de error, renderizar la plantilla con datos vacíos para no romper la UI
        fecha_actual = datetime.now()
        año_actual = fecha_actual.year
        meses_disponibles = get_meses_del_año(año_actual)
        linea_seleccionada_nombre = request.args.get('linea_nombre', 'PETMEDICA')
        lineas_disponibles = [
            'PETMEDICA', 'AGROVET', 'PET NUTRISCIENCE', 'AVIVET', 'OTROS', 'TERCEROS', 'INTERPET'
        ]
        dia_actual = fecha_actual.day
        kpis_default = {
            'meta_total': 0, 'venta_total': 0, 'porcentaje_avance': 0,
            'meta_ipn': 0, 'venta_ipn': 0, 'porcentaje_avance_ipn': 0,
            'vencimiento_6_meses': 0, 'avance_diario_total': 0, 'avance_diario_ipn': 0
        }
        
        return render_template('dashboard_linea.html',
                               linea_nombre=linea_seleccionada_nombre,
                               mes_seleccionado=fecha_actual.strftime('%Y-%m'),
                               meses_disponibles=meses_disponibles,
                               kpis=kpis_default,
                               datos_vendedores=[],
                               datos_productos=[],
                               datos_ciclo_vida=[],
                               datos_forma_farmaceutica=[],
                               lineas_disponibles=lineas_disponibles,
                               fecha_actual=fecha_actual,
                               dia_actual=dia_actual,
                               avance_lineal_pct=0,
                               faltante_meta=0,
                               avance_lineal_ipn_pct=0,
                               faltante_meta_ipn=0,
                               is_admin=is_admin) # Pasar el flag a la plantilla


@app.route('/meta', methods=['GET', 'POST'])
def meta():
    """Ruta deshabilitada - plantilla no existe en este proyecto"""
    if 'username' not in session:
        return redirect(url_for('login'))
    flash('Esta funcionalidad no está disponible en este proyecto.', 'warning')
    return redirect(url_for('dashboard'))



# FUNCIONALIDAD DESHABILITADA - Las plantillas meta.html y metas_vendedor.html no existen
# Esta función fue comentada porque el proyecto clonado solo necesita la conexión a Odoo y ventas
"""
@app.route('/metas_vendedor_DISABLED', methods=['GET', 'POST'])
def metas_vendedor_disabled():
    if 'username' not in session:
        return redirect(url_for('login'))

    # --- Verificación de Permisos ---
    is_admin = session.get('username') in ADMIN_USERS
    if not is_admin:
        flash('No tienes permiso para acceder a esta página.', 'warning')
        return redirect(url_for('dashboard'))
    # Código de la función original comentado...
"""

@app.route('/metas_vendedor', methods=['GET', 'POST'])
def metas_vendedor():
    """Ruta deshabilitada - plantilla no existe en este proyecto"""
    if 'username' not in session:
        return redirect(url_for('login'))
    flash('Esta funcionalidad no está disponible en este proyecto.', 'warning')
    return redirect(url_for('dashboard'))

@app.route('/export/dashboard/details')
def export_dashboard_details():
    """Exporta los detalles del dashboard a un archivo Excel formateado."""
    if 'username' not in session:
        return redirect(url_for('login'))

    # --- Verificación de Permisos ---
    is_admin = session.get('username') in ADMIN_USERS
    if not is_admin:
        flash('No tienes permiso para realizar esta acción.', 'warning')
        return redirect(url_for('dashboard'))
    # --- Fin Verificación ---

    try:
        # Obtener el mes seleccionado de los parámetros de la URL
        mes_seleccionado = request.args.get('mes')
        if not mes_seleccionado:
            flash('No se especificó un mes para la exportación.', 'danger')
            return redirect(url_for('dashboard'))

        # --- Lógica de Fechas (incluyendo filtro de día) ---
        año_sel, mes_sel = mes_seleccionado.split('-')
        fecha_inicio = f"{año_sel}-{mes_sel}-01"

        # Usar el día del parámetro si está disponible, si no, el último día del mes
        dia_fin_param = request.args.get('dia_fin')
        if dia_fin_param and dia_fin_param.isdigit():
            dia_fin = int(dia_fin_param)
            fecha_fin = f"{año_sel}-{mes_sel}-{str(dia_fin).zfill(2)}"
        else:
            # Comportamiento por defecto: mes completo
            ultimo_dia = calendar.monthrange(int(año_sel), int(mes_sel))[1]
            fecha_fin = f"{año_sel}-{mes_sel}-{ultimo_dia}"

        # Obtener datos de ventas reales desde Odoo para ese mes
        sales_data = data_manager.get_sales_lines(
            date_from=fecha_inicio,
            date_to=fecha_fin,
            limit=10000  # Límite alto para exportación
        )

        # Filtrar VENTA INTERNACIONAL (exportaciones), igual que en el dashboard
        sales_data_filtered = []
        for sale in sales_data:
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                if 'VENTA INTERNACIONAL' in linea_comercial[1].upper():
                    continue
            
            canal_ventas = sale.get('sales_channel_id')
            if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                nombre_canal = canal_ventas[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_canal or 'INTERNACIONAL' in nombre_canal:
                    continue
            
            sales_data_filtered.append(sale)

        # --- Procesar datos para un formato legible en Excel ---
        processed_for_excel = []
        for record in sales_data_filtered:
            processed_record = {}
            for key, value in record.items():
                # Si el valor es una lista como [id, 'nombre'], extrae solo el nombre
                if isinstance(value, list) and len(value) > 1:
                    processed_record[key] = value[1]
                else:
                    processed_record[key] = value
            
            # Asegurar que el balance sea un número para el formato de moneda
            if 'balance' in processed_record:
                try:
                    processed_record['balance'] = float(processed_record['balance'])
                except (ValueError, TypeError):
                    processed_record['balance'] = 0.0
            
            processed_for_excel.append(processed_record)

        # Crear DataFrame de Pandas con los datos ya procesados
        df = pd.DataFrame(processed_for_excel)

        # --- TRADUCCIÓN Y ORDEN DE COLUMNAS ---
        column_translations = {
            'invoice_date': 'Fecha Factura',
            'l10n_latam_document_type_id': 'Tipo Documento',
            'move_name': 'Número Documento',
            'partner_name': 'Cliente',
            'vat': 'RUC/DNI Cliente',
            'invoice_user_id': 'Vendedor',
            'default_code': 'Código Producto',
            'name': 'Descripción Producto',
            'quantity': 'Cantidad',
            'price_unit': 'Precio Unitario',
            'balance': 'Importe Total',
            'commercial_line_national_id': 'Línea Comercial',
            'sales_channel_id': 'Canal de Venta',
            'payment_state': 'Estado de Pago',
            'invoice_origin': 'Documento Origen',
            'product_life_cycle': 'Ciclo de Vida Producto',
            'pharmacological_classification_id': 'Clasificación Farmacológica',
            'pharmaceutical_forms_id': 'Forma Farmacéutica',
            'administration_way_id': 'Vía de Administración',
            'production_line_id': 'Línea de Producción',
            'categ_id': 'Categoría de Producto',
            'route_id': 'Ruta de Venta'
        }

        # Filtrar el DataFrame para mantener solo las columnas que vamos a usar
        df = df[list(column_translations.keys())]

        # Renombrar las columnas
        df.rename(columns=column_translations, inplace=True)
        
        # El orden de las columnas en el Excel será el mismo que en el diccionario
        # --- FIN DE TRADUCCIÓN Y ORDEN ---

        # --- Creación y Formateo del Archivo Excel ---
        # Crear archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = f'Detalle Ventas {mes_seleccionado}'
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Obtener el workbook y la worksheet para aplicar estilos
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # --- Definir Estilos ---
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="875A7B", end_color="875A7B", fill_type="solid")
            currency_format = 'S/ #,##0.00;[Red]-S/ #,##0.00'
            date_format = 'YYYY-MM-DD'
            number_format = '#,##0'

            # --- Aplicar Estilos al Encabezado ---
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill

            # --- Aplicar Formato a Columnas y Ajustar Ancho ---
            for col_idx, column in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                
                # Encontrar el ancho máximo
                if len(df[column]) > 0:
                    max_length = max(df[column].astype(str).map(len).max(), len(column)) + 2
                else:
                    max_length = len(column) + 2
                
                worksheet.column_dimensions[col_letter].width = max_length

                # Aplicar formato a celdas específicas
                if column.lower() == 'balance':
                    for cell in worksheet[col_letter][1:]:
                        cell.number_format = currency_format
                elif column.lower() == 'price_unit':
                    for cell in worksheet[col_letter][1:]:
                        cell.number_format = currency_format
                elif column.lower() == 'quantity':
                    for cell in worksheet[col_letter][1:]:
                        cell.number_format = number_format
                elif 'date' in column.lower():
                    for cell in worksheet[col_letter][1:]:
                        # Convertir texto a objeto datetime si es necesario
                        if isinstance(cell.value, str):
                            try:
                                cell.value = datetime.strptime(cell.value, '%Y-%m-%d')
                            except ValueError:
                                pass # Dejar como texto si no se puede convertir
                        cell.number_format = date_format

            # --- Congelar Panel Superior ---
            worksheet.freeze_panes = 'A2'

        # Mover el cursor al inicio del stream para enviarlo
        output.seek(0)
        # --- Fin del Formateo ---

        # Generar nombre de archivo
        filename = f'detalle_ventas_{mes_seleccionado}.xlsx'

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        flash(f'Error al exportar los detalles del dashboard: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))


@app.route('/api/mapa-ventas', methods=['GET'])
def api_mapa_ventas():
    """API endpoint para obtener datos del mapa geográfico de ventas"""
    if 'username' not in session:
        return {'error': 'No autenticado'}, 401
    
    try:
        año = int(request.args.get('año', datetime.now().year))
        mes = int(request.args.get('mes', datetime.now().month))
        canal_filtro = request.args.get('canal', '').upper()  # DIGITAL, NACIONAL, OTROS o '' para todos
        
        # Construir rango de fechas
        fecha_inicio = datetime(año, mes, 1).strftime('%Y-%m-%d')
        ultimo_dia = calendar.monthrange(año, mes)[1]
        fecha_fin = datetime(año, mes, ultimo_dia).strftime('%Y-%m-%d')
        
        # Determinar fuente de datos (meses cerrados de 2026+ -> Supabase)
        source = get_data_source(año, mes)

        if source == 'supabase':
            print(f"🗺️ Obteniendo datos del mapa desde Supabase ({año}-{mes:02d})")
            sales_data = supabase_manager.get_sales_data(fecha_inicio, fecha_fin)
        else:
            print(f"🗺️ Obteniendo datos del mapa desde Odoo ({año}-{mes:02d})")
            sales_data = data_manager.get_sales_lines(date_from=fecha_inicio, date_to=fecha_fin, limit=SALES_LIMIT)
        
        # Validar que sales_data sea una lista
        if not isinstance(sales_data, list):
            print(f"⚠️ sales_data no es una lista: {type(sales_data)}")
            sales_data = []
        
        print(f"📊 Total registros obtenidos para mapa: {len(sales_data)}")
        
        if len(sales_data) == 0:
            return {
                'success': True,
                'data': [],
                'periodo': f"{año}-{mes:02d}",
                'fuente': source,
                'total_provincias': 0,
                'total_ventas': 0
            }
        
        # Mapear canales si hay filtro activo
        cliente_canal_map = {}
        if canal_filtro:
            if source == 'supabase':
                # Para Supabase: Usar campo 'canal' directamente
                for sale in sales_data:
                    canal_directo = sale.get('canal', '')
                    if not canal_directo:
                        continue
                    canal_upper = str(canal_directo).upper()
                    
                    partner_name = sale.get('partner_name', '')
                    if not partner_name:
                        continue
                    
                    # Clasificar según nombre del canal
                    if 'ECOMMERCE' in canal_upper or 'AIRBNB' in canal_upper or 'EMPLEADO' in canal_upper:
                        cliente_canal_map[partner_name] = 'DIGITAL'
                    else:
                        cliente_canal_map[partner_name] = 'NACIONAL'
            else:
                # Para Odoo: Query groups_ids desde res.partner y consultar nombres en agr.groups
                partner_ids = list(set([s.get('partner_id', [0])[0] if isinstance(s.get('partner_id'), list) else s.get('partner_id', 0) for s in sales_data]))
                partner_ids = [pid for pid in partner_ids if pid and pid != 0]
                
                if partner_ids:
                    try:
                        # Obtener partners con sus grupos
                        partners = data_manager.models.execute_kw(
                            data_manager.db, data_manager.uid, data_manager.password,
                            'res.partner', 'search_read',
                            [[['id', 'in', partner_ids]]],
                            {'fields': ['id', 'name', 'groups_ids']}
                        )
                        
                        # Obtener IDs únicos de grupos
                        all_group_ids = set()
                        for partner in partners:
                            groups = partner.get('groups_ids', [])
                            if groups:
                                all_group_ids.update(groups)
                        
                        # Consultar nombres de grupos desde agr.groups
                        group_names = {}
                        if all_group_ids:
                            groups_data = data_manager.models.execute_kw(
                                data_manager.db, data_manager.uid, data_manager.password,
                                'agr.groups', 'search_read',
                                [[['id', 'in', list(all_group_ids)]]],
                                {'fields': ['id', 'name']}
                            )
                            group_names = {g['id']: g['name'].upper() for g in groups_data}
                        
                        # Crear mapeo tanto por ID como por nombre
                        for partner in partners:
                            partner_id = partner.get('id')
                            partner_name = partner.get('name', '')
                            groups_ids = partner.get('groups_ids', [])
                            
                            canal = 'OTROS'  # Por defecto
                            
                            if groups_ids:
                                # Obtener el nombre del primer grupo
                                first_group_id = groups_ids[0]
                                group_name = group_names.get(first_group_id, '')
                                
                                # Clasificar según nombre del grupo
                                if 'ECOMMERCE' in group_name or 'AIRBNB' in group_name or 'EMPLEADO' in group_name:
                                    canal = 'DIGITAL'
                                else:
                                    canal = 'NACIONAL'
                            else:
                                # Sin grupo asignado
                                canal = 'OTROS'
                            
                            # Mapear tanto por ID como por nombre
                            cliente_canal_map[partner_id] = canal
                            cliente_canal_map[partner_name] = canal
                        
                        print(f"🔍 Filtro de canal: {canal_filtro}")
                        print(f"📊 Canales mapeados: {len(cliente_canal_map)} entradas")
                        # Contar por tipo
                        digital = sum(1 for v in cliente_canal_map.values() if v == 'DIGITAL')
                        nacional = sum(1 for v in cliente_canal_map.values() if v == 'NACIONAL')
                        otros = sum(1 for v in cliente_canal_map.values() if v == 'OTROS')
                        print(f"   - DIGITAL: {digital} clientes")
                        print(f"   - NACIONAL: {nacional} clientes")
                        print(f"   - OTROS: {otros} clientes")
                    except Exception as e:
                        print(f"⚠️ Error obteniendo canales para mapa: {e}")
        
        # Procesar datos por provincia
        ventas_por_provincia = {}
        clientes_por_provincia = {}
        registros_sin_provincia = 0
        registros_con_provincia = 0
        
        for idx, sale in enumerate(sales_data):
            try:
                # Validar que sale sea un diccionario
                if not isinstance(sale, dict):
                    print(f"⚠️ Registro {idx} no es dict: {type(sale)}")
                    continue
                
                # Aplicar filtro de canal si está activo
                if canal_filtro:
                    # Obtener partner_name e partner_id
                    partner_info = sale.get('partner_id')
                    partner_name = sale.get('partner_name', '')
                    
                    # Extraer ID si partner_id es lista
                    if isinstance(partner_info, list) and len(partner_info) > 0:
                        partner_id = partner_info[0]
                        if not partner_name and len(partner_info) > 1:
                            partner_name = partner_info[1]
                    else:
                        partner_id = partner_info
                    
                    # Buscar canal por ID primero, luego por nombre
                    canal_cliente = cliente_canal_map.get(partner_id) or cliente_canal_map.get(partner_name, 'OTROS')
                    
                    if canal_cliente != canal_filtro:
                        continue
                
                # Obtener provincia (state_id o state_name)
                provincia_info = sale.get('state_id') or sale.get('provincia') or sale.get('state_name')
                if not provincia_info:
                    registros_sin_provincia += 1
                    continue
                
                # Normalizar nombre de provincia
                if isinstance(provincia_info, list) and len(provincia_info) > 1:
                    provincia_nombre = str(provincia_info[1]).upper()
                elif isinstance(provincia_info, str):
                    provincia_nombre = provincia_info.upper()
                else:
                    # Si provincia_info es un número, intentar obtener state_name
                    if sale.get('state_name'):
                        provincia_nombre = str(sale.get('state_name')).upper()
                    else:
                        registros_sin_provincia += 1
                        continue
                
                registros_con_provincia += 1
                
                # Limpiar nombre
                provincia_nombre = re.sub(r'\s*\(PE\)\s*', '', provincia_nombre, flags=re.IGNORECASE).strip()
                provincia_nombre = provincia_nombre.replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')
                
                # Mapeos específicos
                if 'CALLAO' in provincia_nombre:
                    provincia_nombre = 'CALLAO'
                elif 'MARTIN' in provincia_nombre:
                    provincia_nombre = 'SAN MARTIN'
                
                # Obtener monto (manejar tanto 'balance' como 'price_subtotal')
                balance = sale.get('balance') or sale.get('price_subtotal', 0)
                
                # Convertir a float de manera segura
                if isinstance(balance, str):
                    balance = float(balance.replace(',', '').replace('S/', '').strip())
                elif balance is None:
                    balance = 0
                else:
                    balance = float(balance)
                
                # Acumular ventas
                ventas_por_provincia[provincia_nombre] = ventas_por_provincia.get(provincia_nombre, 0) + balance
                
                # Contar clientes únicos
                partner_id = sale.get('partner_id')
                if partner_id:
                    if isinstance(partner_id, list) and len(partner_id) > 0:
                        partner_id = partner_id[0]
                    
                    if provincia_nombre not in clientes_por_provincia:
                        clientes_por_provincia[provincia_nombre] = set()
                    clientes_por_provincia[provincia_nombre].add(partner_id)
            
            except Exception as e:
                print(f"⚠️ Error procesando registro {idx}: {e}")
                continue
        
        # Preparar respuesta
        mapa_data = []
        for provincia, ventas in ventas_por_provincia.items():
            num_clientes = len(clientes_por_provincia.get(provincia, set()))
            mapa_data.append({
                'name': provincia,
                'value': round(ventas, 2),
                'clientes': num_clientes,
                'ticket_promedio': round(ventas / num_clientes, 2) if num_clientes > 0 else 0
            })
        
        print(f"   - Registros con provincia: {registros_con_provincia}")
        print(f"   - Registros sin provincia: {registros_sin_provincia}")
        if len(mapa_data) > 0:
            print(f"   - Top 3 provincias: {[d['name'] for d in mapa_data[:3]]}")
        # Ordenar por ventas descendente
        mapa_data.sort(key=lambda x: x['value'], reverse=True)
        
        print(f"🗺️ Mapa generado: {len(mapa_data)} provincias con ventas")
        
        return {
            'success': True,
            'data': mapa_data,
            'periodo': f"{año}-{mes:02d}",
            'fuente': source,
            'total_provincias': len(mapa_data),
            'total_ventas': sum(ventas_por_provincia.values())
        }
    
    except Exception as e:
        print(f"❌ Error en API mapa-ventas: {e}")
        return {'error': str(e)}, 500


@app.route('/api/cobertura-filtrada', methods=['GET'])
def api_cobertura_filtrada():
    """API endpoint para obtener cobertura de clientes filtrada por canal de venta"""
    if 'username' not in session:
        return {'error': 'No autenticado'}, 401
    
    try:
        # Obtener parámetros
        mes_str = request.args.get('mes', datetime.now().strftime('%Y-%m'))
        año = int(request.args.get('año', datetime.now().year))
        canal_filtro = request.args.get('canal', 'TODOS')
        
        print(f"🔍 API Cobertura Filtrada por Canal: mes={mes_str}, año={año}, canal={canal_filtro}")
        
        # Parsear mes
        año_mes, mes_num = mes_str.split('-')
        mes_int = int(mes_num)
        año_int = int(año_mes)
        
        # Construir fechas
        fecha_inicio = datetime(año_int, mes_int, 1)
        ultimo_dia = calendar.monthrange(año_int, mes_int)[1]
        fecha_fin = datetime(año_int, mes_int, ultimo_dia)
        fecha_inicio_ano = datetime(año_int, 1, 1)
        
        # Determinar fuente de datos (meses cerrados de 2026+ -> Supabase)
        source = get_data_source(año_int, mes_int)

        if source != 'odoo':
            # Para Supabase, calcular cobertura usando campo 'canal'
            print(f"📊 Año {año_int} usa Supabase - calculando cobertura por canal")
            
            # Obtener clientes activos y cartera desde Supabase
            fecha_inicio_str = fecha_inicio.strftime('%Y-%m-%d')
            fecha_fin_str = fecha_fin.strftime('%Y-%m-%d')
            fecha_inicio_ano_str = fecha_inicio_ano.strftime('%Y-%m-%d')
            
            # Cartera del año
            ventas_ano = supabase_manager.get_sales_data(fecha_inicio_ano_str, fecha_fin_str)
            cartera_ids = set(v.get('partner_id') for v in ventas_ano if v.get('partner_id'))
            
            # Activos del mes
            ventas_mes = supabase_manager.get_sales_data(fecha_inicio_str, fecha_fin_str)
            activos_ids = set(v.get('partner_id') for v in ventas_mes if v.get('partner_id'))
            
            # Agrupar por canal
            cartera_por_canal = {}
            activos_por_canal = {}
            
            for venta in ventas_ano:
                canal = venta.get('canal', 'SIN CANAL')
                partner_id = venta.get('partner_id')
                if not partner_id:
                    continue
                
                # Clasificar en DIGITAL o NACIONAL
                canal_upper = str(canal).upper()
                if 'ECOMMERCE' in canal_upper or 'AIRBNB' in canal_upper or 'EMPLEADO' in canal_upper:
                    grupo_clasificado = 'DIGITAL'
                else:
                    grupo_clasificado = 'NACIONAL'
                
                # Aplicar filtro
                if canal_filtro != 'TODOS' and grupo_clasificado != canal_filtro:
                    continue
                
                if grupo_clasificado not in cartera_por_canal:
                    cartera_por_canal[grupo_clasificado] = set()
                cartera_por_canal[grupo_clasificado].add(partner_id)
            
            for venta in ventas_mes:
                canal = venta.get('canal', 'SIN CANAL')
                partner_id = venta.get('partner_id')
                if not partner_id:
                    continue
                
                # Clasificar en DIGITAL o NACIONAL
                canal_upper = str(canal).upper()
                if 'ECOMMERCE' in canal_upper or 'AIRBNB' in canal_upper or 'EMPLEADO' in canal_upper:
                    grupo_clasificado = 'DIGITAL'
                else:
                    grupo_clasificado = 'NACIONAL'
                
                # Aplicar filtro
                if canal_filtro != 'TODOS' and grupo_clasificado != canal_filtro:
                    continue
                
                if grupo_clasificado not in activos_por_canal:
                    activos_por_canal[grupo_clasificado] = set()
                activos_por_canal[grupo_clasificado].add(partner_id)
            
            # Construir respuesta
            datos_grupos = []
            total_cartera_global = 0
            total_activos_global = 0
            
            for grupo in sorted(cartera_por_canal.keys()):
                cartera_count = len(cartera_por_canal.get(grupo, set()))
                activos_count = len(activos_por_canal.get(grupo, set()))
                cobertura = (activos_count / cartera_count * 100) if cartera_count > 0 else 0
                
                datos_grupos.append({
                    'grupo': grupo,
                    'cartera': cartera_count,
                    'activos': activos_count,
                    'cobertura': cobertura
                })
                
                total_cartera_global += cartera_count
                total_activos_global += activos_count
            
            # Agregar total
            cobertura_general = (total_activos_global / total_cartera_global * 100) if total_cartera_global > 0 else 0
            
            datos_grupos.append({
                'grupo': 'TOTAL GENERAL',
                'cartera': total_cartera_global,
                'activos': total_activos_global,
                'cobertura': cobertura_general,
                'es_total': True
            })
            
            print(f"✅ Cobertura Supabase calculada: {total_activos_global}/{total_cartera_global} = {cobertura_general:.1f}%")
            
            return {
                'success': True,
                'cobertura': round(cobertura_general, 1),
                'grupos': datos_grupos,
                'canal': canal_filtro
            }
        
        # Obtener todos los grupos
        grupos_data = data_manager.models.execute_kw(
            data_manager.db, data_manager.uid, data_manager.password,
            'agr.groups', 'search_read',
            [[]],
            {'fields': ['id', 'name']}
        )
        grupos_dict = {g['id']: g['name'] for g in grupos_data}
        
        print(f"   📋 Grupos encontrados: {len(grupos_dict)}")
        
        # Calcular cobertura por grupo
        datos_grupos = []
        total_cartera_global = 0
        total_activos_global = 0
        
        # Filtrar grupos según canal (DIGITAL = ECOMMERCE, AIRBNB, EMPLEADOS | NACIONAL = resto)
        grupos_digitales = ['ECOMMERCE', 'AIRBNB', 'EMPLEADOS']
        
        for grupo_id, grupo_nombre in grupos_dict.items():
            # Aplicar filtro de canal
            if canal_filtro == 'DIGITAL' and grupo_nombre.upper() not in grupos_digitales:
                continue
            elif canal_filtro == 'NACIONAL' and grupo_nombre.upper() in grupos_digitales:
                continue
            
            # Obtener partners con este grupo
            domain_partners = [('groups_ids', 'in', [grupo_id]), ('customer_rank', '>', 0)]
            
            partners_ids = data_manager.models.execute_kw(
                data_manager.db, data_manager.uid, data_manager.password,
                'res.partner', 'search', [domain_partners]
            )
            
            if not partners_ids:
                continue
            
            print(f"   🔍 Grupo {grupo_nombre}: {len(partners_ids)} partners")
            
            # Cartera: partners del grupo que compraron en el año
            domain_cartera = [
                ('partner_id', 'in', partners_ids),
                ('move_id.state', '=', 'posted'),
                ('move_id.move_type', 'in', ['out_invoice', 'out_refund']),
                ('date', '>=', fecha_inicio_ano.strftime('%Y-%m-%d')),
                ('date', '<=', fecha_fin.strftime('%Y-%m-%d')),
                ('product_id', '!=', False)
            ]
            
            cartera_result = data_manager.models.execute_kw(
                data_manager.db, data_manager.uid, data_manager.password,
                'account.move.line', 'read_group',
                [domain_cartera],
                {'fields': ['partner_id'], 'groupby': ['partner_id'], 'lazy': False}
            )
            total_cartera = len(cartera_result)
            
            # Activos: partners del grupo que compraron en el mes
            domain_activos = [
                ('partner_id', 'in', partners_ids),
                ('move_id.state', '=', 'posted'),
                ('move_id.move_type', 'in', ['out_invoice', 'out_refund']),
                ('date', '>=', fecha_inicio.strftime('%Y-%m-%d')),
                ('date', '<=', fecha_fin.strftime('%Y-%m-%d')),
                ('product_id', '!=', False)
            ]
            
            activos_result = data_manager.models.execute_kw(
                data_manager.db, data_manager.uid, data_manager.password,
                'account.move.line', 'read_group',
                [domain_activos],
                {'fields': ['partner_id'], 'groupby': ['partner_id'], 'lazy': False}
            )
            total_activos = len(activos_result)
            
            cobertura_grupo = (total_activos / total_cartera * 100) if total_cartera > 0 else 0
            
            datos_grupos.append({
                'grupo': grupo_nombre,
                'cartera': total_cartera,
                'activos': total_activos,
                'cobertura': cobertura_grupo
            })
            
            total_cartera_global += total_cartera
            total_activos_global += total_activos
        
        # Agregar total
        datos_grupos.append({
            'grupo': 'TOTAL GENERAL',
            'cartera': total_cartera_global,
            'activos': total_activos_global,
            'cobertura': (total_activos_global / total_cartera_global * 100) if total_cartera_global > 0 else 0,
            'es_total': True
        })
        
        # Calcular cobertura general para el gauge
        cobertura_general = (total_activos_global / total_cartera_global * 100) if total_cartera_global > 0 else 0
        
        print(f"✅ Cobertura calculada: {total_activos_global}/{total_cartera_global} = {cobertura_general:.1f}%")
        
        return {
            'success': True,
            'cobertura': round(cobertura_general, 1),
            'grupos': datos_grupos,
            'canal': canal_filtro
        }
    
    except Exception as e:
        print(f"❌ Error en API cobertura-filtrada: {e}")
        import traceback
        traceback.print_exc()
        return {'error': str(e)}, 500


if __name__ == '__main__':
    # Soporte para Render.com: usar puerto dinámico
    port = int(os.environ.get("PORT", 5000))
    debug_mode = os.environ.get("FLASK_ENV") != "production"
    
    print("🚀 Iniciando Dashboard de Ventas Farmacéuticas...")
    print(f"📊 Puerto: {port}")
    print(f"🔧 Modo debug: {debug_mode}")
    print("🔐 Usuario: configurado en variables de entorno")
    
    app.run(host='0.0.0.0', port=port, debug=debug_mode)
